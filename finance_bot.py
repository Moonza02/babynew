"""BabyDiary moliya boti — BITTA FAYL.
Asosiy bot bilan bir Railway service'da, launcher.py orqali ishga tushadi.
Ishga tushirish: python finance_bot.py
"""
import sys as _sys
# Barcha ichki "modul" nomlari shu bitta faylga ishora qiladi:
config = db = store = finance = report_pdf = report_xlsx = jobs = handlers = _sys.modules[__name__]
# ===================== config.py =====================
"""BabyDiary moliya boti — sozlamalar (asosiy bot bilan BIR service)."""
import os
from zoneinfo import ZoneInfo

FINANCE_BOT_TOKEN = os.getenv("FINANCE_BOT_TOKEN", "")

# Investorlar (faqat shu 2 kishi)
INVESTORS = {
    5285940949: "Ibrohim",
    512101064:  "Rustam",
}
ADMIN_IDS = set(INVESTORS.keys())

# Umumiy ma'lumot papkasi (asosiy bot ham shu /data ga yozadi)
DATA_DIR = "/data" if os.path.isdir("/data") else os.path.join(os.path.dirname(__file__), "data")

# Moliya botining O'Z bazasi
FINANCE_DB = os.path.join(DATA_DIR, "finance.db")

# Asosiy bot yozadigan JONLI fayllar — to'g'ridan-to'g'ri o'qiymiz (real-time)
BABYDIARY_DB  = os.getenv("BABYDIARY_DB",  os.path.join(DATA_DIR, "babydiary.db"))
ORDERS_JSON   = os.getenv("ORDERS_JSON",   os.path.join(DATA_DIR, "orders.json"))
PRODUCTS_JSON = os.getenv("PRODUCTS_JSON", os.path.join(DATA_DIR, "products.json"))

# Hisobot / backup
REPORT_CHAT_IDS   = [int(x) for x in os.getenv("REPORT_CHAT_IDS", "").split(",") if x] or list(ADMIN_IDS)
BACKUP_CHANNEL_ID = int(os.getenv("BACKUP_CHANNEL_ID", "0")) or None

# Daromad qaysi statusda hisoblansin (babydiary.db order_tracking bo'yicha)
REVENUE_STATUSES = ("confirmed", "preparing", "delivering", "delivered")

TZ = ZoneInfo("Asia/Tashkent")
os.makedirs(DATA_DIR, exist_ok=True)


# ===================== db.py =====================
"""finance.db — rasxod, tenglashtirish tarixi, sozlamalar."""
import sqlite3
from datetime import datetime
from contextlib import contextmanager


def init_db():
    with sqlite3.connect(config.FINANCE_DB) as con:
        con.execute("PRAGMA journal_mode=WAL;")
        con.executescript("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount INTEGER NOT NULL,
            note TEXT,
            kind TEXT NOT NULL DEFAULT 'operatsion',  -- 'tovar' | 'operatsion'
            payer_id INTEGER,
            payer_name TEXT,
            photo_id TEXT,          -- chek/rasm (ixtiyoriy)
            del_req_by INTEGER,     -- o'chirishni so'ragan investor (0/None = so'rov yo'q)
            subcat TEXT,            -- operatsion kichik kategoriyasi
            edit_req_by INTEGER,    -- tahrirlashni so'ragan investor
            edit_amount INTEGER,    -- taklif qilingan yangi summa
            edit_note TEXT,         -- taklif qilingan yangi izoh
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS settlements (
            period TEXT PRIMARY KEY,           -- 'YYYY-MM'
            payer TEXT, receiver TEXT,
            amount INTEGER,
            paid INTEGER DEFAULT 0,            -- 0=to'lanmagan, 1=to'langan
            created_at TEXT
        );
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY, value TEXT
        );
        CREATE TABLE IF NOT EXISTS report_log (
            period TEXT PRIMARY KEY, sent_at TEXT
        );
        """)
        # Migratsiya — eski bazada ustun bo'lmasa qo'shamiz
        for col, ddl in (("photo_id", "TEXT"), ("del_req_by", "INTEGER"),
                         ("subcat", "TEXT"), ("edit_req_by", "INTEGER"),
                         ("edit_amount", "INTEGER"), ("edit_note", "TEXT")):
            try:
                con.execute(f"ALTER TABLE expenses ADD COLUMN {col} {ddl}")
            except Exception:
                pass


@contextmanager
def conn():
    c = sqlite3.connect(config.FINANCE_DB)
    c.row_factory = sqlite3.Row
    try:
        yield c; c.commit()
    finally:
        c.close()


def _now():
    return datetime.now(config.TZ).isoformat()


# ---- Sozlamalar ----
def get_setting(key, default=None):
    with conn() as c:
        r = c.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return r["value"] if r else default

def set_setting(key, value):
    with conn() as c:
        c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, str(value)))

def gift_box_cost():
    return int(get_setting("gift_box_cost", "0") or 0)

def qadoq_oddiy_cost():
    """Oddiy (sovg'a qutisisiz) buyurtma uchun to'liq qadoq tannarxi:
    tissue + vizitka + thank you card + nakleyka + sumka."""
    return int(get_setting("qadoq_oddiy_cost", "0") or 0)


# ---- Rasxod ----
def add_expense(amount, note, kind, payer_id, payer_name, photo_id=None, subcat=""):
    with conn() as c:
        cur = c.execute(
            "INSERT INTO expenses(amount,note,kind,payer_id,payer_name,photo_id,subcat,created_at) "
            "VALUES(?,?,?,?,?,?,?,?)",
            (amount, note, kind, payer_id, payer_name, photo_id, subcat, _now()))
        return cur.lastrowid

def get_expense(rid):
    with conn() as c:
        r = c.execute("SELECT * FROM expenses WHERE id=?", (rid,)).fetchone()
        return dict(r) if r else None

def delete_expense(rid):
    with conn() as c:
        return c.execute("DELETE FROM expenses WHERE id=?", (rid,)).rowcount > 0

def request_delete(rid, by_id):
    """O'chirish so'rovini belgilaydi (ikkinchi investor tasdiqlashi kerak)."""
    with conn() as c:
        return c.execute("UPDATE expenses SET del_req_by=? WHERE id=?", (by_id, rid)).rowcount > 0

def cancel_delete(rid):
    with conn() as c:
        return c.execute("UPDATE expenses SET del_req_by=NULL WHERE id=?", (rid,)).rowcount > 0

def request_edit(rid, by_id, amount, note):
    """Tahrirlash so'rovi — yangi summa/izoh saqlanadi, ikkinchi investor tasdiqlaydi."""
    with conn() as c:
        return c.execute("UPDATE expenses SET edit_req_by=?, edit_amount=?, edit_note=? WHERE id=?",
                         (by_id, int(amount), note, rid)).rowcount > 0

def cancel_edit(rid):
    with conn() as c:
        return c.execute("UPDATE expenses SET edit_req_by=NULL, edit_amount=NULL, edit_note=NULL WHERE id=?",
                         (rid,)).rowcount > 0

def apply_edit(rid):
    """Tasdiqlangan tahrirni qo'llaydi."""
    with conn() as c:
        r = c.execute("SELECT edit_req_by, edit_amount, edit_note FROM expenses WHERE id=?", (rid,)).fetchone()
        if not r or r["edit_req_by"] is None:
            return False
        c.execute("UPDATE expenses SET amount=?, note=?, edit_req_by=NULL, edit_amount=NULL, edit_note=NULL WHERE id=?",
                  (int(r["edit_amount"]), r["edit_note"], rid))
        return True

def month_expenses(year, month):
    ym = f"{year:04d}-{month:02d}"
    with conn() as c:
        rows = c.execute("SELECT * FROM expenses WHERE substr(created_at,1,7)=? ORDER BY id", (ym,)).fetchall()
    return [dict(r) for r in rows]

def range_expenses(start, end):
    """start..end (YYYY-MM-DD, ikkalasi ham kiritiladi) oralig'idagi rasxodlar."""
    with conn() as c:
        rows = c.execute(
            "SELECT * FROM expenses WHERE substr(created_at,1,10) BETWEEN ? AND ? ORDER BY id",
            (start, end)).fetchall()
    return [dict(r) for r in rows]

def all_expenses():
    with conn() as c:
        return [dict(r) for r in c.execute("SELECT * FROM expenses ORDER BY id").fetchall()]


# ---- Tenglashtirish tarixi ----
def save_settlement(period, payer, receiver, amount):
    with conn() as c:
        c.execute("""INSERT OR REPLACE INTO settlements(period,payer,receiver,amount,paid,created_at)
                     VALUES(?,?,?,?,COALESCE((SELECT paid FROM settlements WHERE period=?),0),?)""",
                  (period, payer, receiver, int(amount), period, _now()))

def mark_settlement_paid(period, paid=1):
    with conn() as c:
        return c.execute("UPDATE settlements SET paid=? WHERE period=?", (paid, period)).rowcount > 0

def get_settlements():
    with conn() as c:
        return [dict(r) for r in c.execute("SELECT * FROM settlements ORDER BY period DESC").fetchall()]

def unpaid_settlement_balance():
    """To'lanmagan tenglashtirishlar — qarama-qarshi yo'nalishlar NETTO qilinadi.
    Masalan: Rustam→Ibrohim 830k va Ibrohim→Rustam 34k bo'lsa,
    yakuniy: Rustam→Ibrohim 796k (bittagina yo'nalish)."""
    pair = {}  # (sorted_a, sorted_b) -> signed summa (musbat = a→b)
    for s in get_settlements():
        if s["paid"] or not s["amount"]:
            continue
        p, r, amt = s["payer"], s["receiver"], s["amount"]
        a, b = sorted([p, r])
        sign = 1 if p == a else -1
        pair[(a, b)] = pair.get((a, b), 0) + sign * amt
    net = {}
    for (a, b), val in pair.items():
        if val > 0:
            net[(a, b)] = val      # a -> b
        elif val < 0:
            net[(b, a)] = -val     # b -> a
        # val == 0 -> teng, ko'rsatilmaydi
    return net


# ---- report log ----
def report_sent(period):
    with conn() as c:
        return c.execute("SELECT 1 FROM report_log WHERE period=?", (period,)).fetchone() is not None

def mark_report_sent(period):
    with conn() as c:
        c.execute("INSERT OR REPLACE INTO report_log(period,sent_at) VALUES(?,?)", (period, _now()))


# ===================== store.py =====================
"""Asosiy botdan forward qilingan fayllardan savdo ma'lumotini o'qish."""
import os, json, sqlite3, shutil


def save_uploaded(src_path, kind):
    """kind: 'db' | 'orders' | 'products'. Forward qilingan faylni saqlaydi."""
    dst = {"db": config.BABYDIARY_DB, "orders": config.ORDERS_JSON,
           "products": config.PRODUCTS_JSON}[kind]
    shutil.copy2(src_path, dst)
    return dst


def _load_json(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def load_orders():   return _load_json(config.ORDERS_JSON)
def load_products(): return _load_json(config.PRODUCTS_JSON)

def have_orders():   return os.path.exists(config.ORDERS_JSON)
def have_db():       return os.path.exists(config.BABYDIARY_DB)


def product_cost_map():
    """{product_id(str): cost} — products.json'dan."""
    m = {}
    for p in load_products():
        m[str(p.get("id"))] = int(p.get("cost", 0) or 0)
    return m


def order_status_map():
    """{order_id: status} — babydiary.db order_tracking'dan. DB yo'q bo'lsa {}."""
    if not have_db():
        return {}
    try:
        con = sqlite3.connect(config.BABYDIARY_DB)
        rows = con.execute("SELECT order_id, status FROM order_tracking").fetchall()
        con.close()
        return {r[0]: r[1] for r in rows}
    except Exception:
        return {}


def bd_setting(key, default=0):
    """babydiary.db (asosiy bot) settings — qadoq tannarxlari shu yerdan o'rnatiladi."""
    if not have_db():
        return default
    try:
        con = sqlite3.connect(config.BABYDIARY_DB)
        row = con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        con.close()
        return int(row[0]) if row and str(row[0]).strip() != "" else default
    except Exception:
        return default


def inventory_value():
    """Omborda qotgan pul: sum(cost * stock)."""
    total = 0
    for p in load_products():
        total += int(p.get("cost", 0) or 0) * int(p.get("stock", 0) or 0)
    return total


def month_sales(year, month):
    """Shu oy savdo ko'rsatkichlari (range_sales ustiga)."""
    import calendar
    last = calendar.monthrange(year, month)[1]
    return range_sales(f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last:02d}")


def range_sales(start, end):
    """
    start..end (YYYY-MM-DD, ikkalasi kiritiladi) oralig'idagi savdo ko'rsatkichlari.
    Qaytaradi: dict(kirim, delivery, cogs, qadoq_*, orders_count, status_filtered, paytur ...)
    """
    statuses = order_status_map()
    use_status = have_db()
    cost_map = product_cost_map()
    bd_gift  = bd_setting("gift_box_cost", 0)
    bd_oddiy = bd_setting("qadoq_oddiy_cost", 0)

    kirim = delivery = cogs = qadoq_sotuv = 0
    qadoq_soni = orders_count = cashback_used = 0
    qadoq_oddiy_tannarxi = qadoq_quti_tannarxi = 0
    paytur = {}

    for o in load_orders():
        date = (o.get("date") or "")[:10]
        if not date or date < start or date > end:
            continue
        if use_status:
            st = statuses.get(o.get("id"))
            if st not in config.REVENUE_STATUSES:
                continue
        orders_count += 1
        kirim += int(o.get("total", 0) or 0)
        delivery += int(o.get("delivery", 0) or 0)
        cashback_used += int(o.get("cashback_used", 0) or 0)
        pkg_price = int(o.get("packaging_price", 0) or 0)
        qutili = pkg_price > 0
        if qutili:
            qadoq_soni += 1
            qadoq_sotuv += pkg_price
        qc = o.get("qadoq_cost")
        if qc is None:
            qc = bd_gift if qutili else bd_oddiy
        qc = int(qc)
        if qutili:
            qadoq_quti_tannarxi += qc
        else:
            qadoq_oddiy_tannarxi += qc
        for it in o.get("items", []):
            qty = int(it.get("qty", 0) or 0)
            cost = it.get("cost")
            if cost is None:
                cost = cost_map.get(str(it.get("product_id")), 0)
            cogs += int(cost) * qty
        pt = o.get("pay_type", "?")
        paytur[pt] = paytur.get(pt, 0) + int(o.get("total", 0) or 0)

    return {
        "kirim": kirim, "delivery": delivery, "cogs": cogs,
        "qadoq_soni": qadoq_soni, "qadoq_sotuv": qadoq_sotuv,
        "qadoq_oddiy_tannarxi": qadoq_oddiy_tannarxi, "qadoq_quti_tannarxi": qadoq_quti_tannarxi,
        "qadoq_tannarxi": qadoq_oddiy_tannarxi + qadoq_quti_tannarxi,
        "qadoq_oddiy_soni": max(0, orders_count - qadoq_soni), "qadoq_quti_soni": qadoq_soni,
        "orders_count": orders_count, "status_filtered": use_status,
        "cashback_used": cashback_used,
        "paytur": paytur,
    }


def month_product_stats(year, month):
    """Shu oydagi har mahsulot kesimi (range ustiga)."""
    import calendar
    last = calendar.monthrange(year, month)[1]
    return range_product_stats(f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{last:02d}")


def range_product_stats(start, end):
    """start..end oralig'idagi har mahsulot kesimi: [{name, qty, revenue, foyda}]."""
    statuses = order_status_map()
    use_status = have_db()
    cost_map = product_cost_map()
    agg = {}
    for o in load_orders():
        date = (o.get("date") or "")[:10]
        if not date or date < start or date > end:
            continue
        if use_status and statuses.get(o.get("id")) not in config.REVENUE_STATUSES:
            continue
        for it in o.get("items", []):
            name = it.get("name") or it.get("name_uz") or "?"
            qty = int(it.get("qty", 0) or 0)
            price = int(it.get("price", 0) or 0)
            cost = it.get("cost")
            if cost is None:
                cost = cost_map.get(str(it.get("product_id")), 0)
            a = agg.setdefault(name, {"name": name, "qty": 0, "revenue": 0, "foyda": 0})
            a["qty"] += qty
            a["revenue"] += price * qty
            a["foyda"] += (price - int(cost)) * qty
    return list(agg.values())


def month_top_products(year, month, limit=5):
    """Shu oyda eng ko'p sotilgan mahsulotlar (soni bo'yicha)."""
    rows = sorted(month_product_stats(year, month), key=lambda x: x["qty"], reverse=True)
    return rows[:limit]


def product_profit_ranking(year, month, limit=5):
    """Foyda bo'yicha reyting: (eng_foydali[], eng_kam_foydali[])."""
    rows = sorted(month_product_stats(year, month), key=lambda x: x["foyda"], reverse=True)
    top = rows[:limit]
    bottom = list(reversed(rows[-limit:])) if len(rows) > limit else []
    return top, bottom


# ---------- Sotilmayotgan tovar (dead stock) ----------
def last_sold_map():
    """{product_id(str): oxirgi sotilgan sana 'YYYY-MM-DD'}."""
    m = {}
    for o in load_orders():
        date = (o.get("date") or "")[:10]
        for it in o.get("items", []):
            pid = str(it.get("product_id"))
            if pid and pid != "None" and (pid not in m or date > m[pid]):
                m[pid] = date
    return m


def dead_stock(days=30):
    """Omborda turgan (stock>0), lekin `days` kunda sotilmagan tovarlar.
    Qotgan pul (cost*stock) bo'yicha kamayish tartibida."""
    from datetime import timedelta
    cutoff = (datetime.now(config.TZ).date() - timedelta(days=days)).isoformat()
    last = last_sold_map()
    out = []
    for p in load_products():
        stock = int(p.get("stock", 0) or 0)
        if stock <= 0:
            continue
        ls = last.get(str(p.get("id")))
        if ls is None or ls < cutoff:
            out.append({
                "name": p.get("name_uz") or p.get("name") or "?",
                "stock": stock, "cost": int(p.get("cost", 0) or 0), "last": ls,
            })
    out.sort(key=lambda x: x["cost"] * x["stock"], reverse=True)
    return out


# ---------- Mijoz analitikasi ----------
def _norm_phone(p):
    import re
    d = re.sub(r"\D", "", p or "")
    if len(d) == 9:
        d = "998" + d
    elif len(d) == 12 and d.startswith("998"):
        pass
    return d


def customer_stats():
    """Takroriy mijozlar, o'rtacha chek, top mijozlar."""
    statuses = order_status_map()
    use_status = have_db()
    agg = {}
    for o in load_orders():
        if use_status and statuses.get(o.get("id")) not in config.REVENUE_STATUSES:
            continue
        ph = _norm_phone(o.get("phone"))
        if not ph:
            continue
        a = agg.setdefault(ph, {"phone": ph, "name": "", "count": 0, "total": 0})
        a["count"] += 1
        a["total"] += int(o.get("total", 0) or 0)
        if o.get("name"):
            a["name"] = o.get("name")
    custs = list(agg.values())
    total_orders = sum(c["count"] for c in custs)
    total_rev = sum(c["total"] for c in custs)
    repeat = sum(1 for c in custs if c["count"] > 1)
    return {
        "customers": len(custs),
        "orders": total_orders,
        "repeat": repeat,
        "repeat_pct": (repeat / len(custs) * 100) if custs else 0,
        "aov": (total_rev / total_orders) if total_orders else 0,
        "top": sorted(custs, key=lambda x: x["total"], reverse=True)[:5],
    }


# ---------- Promo kod samaradorligi ----------
def promo_stats():
    """Har promo kod: {code, count, discount, revenue}."""
    statuses = order_status_map()
    use_status = have_db()
    agg = {}
    for o in load_orders():
        if use_status and statuses.get(o.get("id")) not in config.REVENUE_STATUSES:
            continue
        code = (o.get("promo_code") or "").strip().upper()
        if not code:
            continue
        a = agg.setdefault(code, {"code": code, "count": 0, "discount": 0, "revenue": 0})
        a["count"] += 1
        a["discount"] += int(o.get("promo_discount", 0) or 0)
        a["revenue"] += int(o.get("total", 0) or 0)
    return sorted(agg.values(), key=lambda x: x["count"], reverse=True)


# ---------- Mijozlarda turgan cashback (majburiyat, foydadan AYRILMAYDI) ----------
def outstanding_cashback():
    if not have_db():
        return 0
    try:
        con = sqlite3.connect(config.BABYDIARY_DB)
        try:
            row = con.execute("SELECT COALESCE(SUM(balance),0) FROM cashback").fetchone()
        except Exception:
            row = con.execute("SELECT COALESCE(SUM(amount),0) FROM cashback").fetchone()
        con.close()
        return int(row[0] or 0) if row else 0
    except Exception:
        return 0


# ===================== finance.py =====================
"""Moliyaviy hisob-kitob: model B (COGS), 50/50, tenglashtirish, balans."""


def somm(n):
    return f"{int(round(n)):,}".replace(",", " ") + " so'm"


def md_escape(t):
    """Telegram Markdown'ni buzadigan belgilarni himoyalaydi (izohlar uchun)."""
    t = str(t or "")
    for ch in ("\\", "_", "*", "`", "["):
        t = t.replace(ch, "\\" + ch)
    return t


def _compute_core(ms, exps, *, period, label, year=None, month=None, start=None, end=None):
    operatsion = sum(e["amount"] for e in exps if e["kind"] == "operatsion")
    tovar      = sum(e["amount"] for e in exps if e["kind"] == "tovar")
    quti_soni  = ms.get("qadoq_quti_soni", ms["qadoq_soni"])
    oddiy_soni = ms.get("qadoq_oddiy_soni", max(0, ms["orders_count"] - ms["qadoq_soni"]))
    qadoq_quti_tannarxi  = ms.get("qadoq_quti_tannarxi", 0)
    qadoq_oddiy_tannarxi = ms.get("qadoq_oddiy_tannarxi", 0)
    qadoq_tannarxi = ms.get("qadoq_tannarxi", qadoq_quti_tannarxi + qadoq_oddiy_tannarxi)

    op_breakdown = {}
    for e in exps:
        if e["kind"] == "operatsion":
            sc = (e.get("subcat") or "Boshqa")
            op_breakdown[sc] = op_breakdown.get(sc, 0) + e["amount"]
    op_breakdown = dict(sorted(op_breakdown.items(), key=lambda x: x[1], reverse=True))

    kirim, delivery, cogs = ms["kirim"], ms["delivery"], ms["cogs"]
    sof_foyda = kirim - delivery - cogs - qadoq_tannarxi - operatsion
    profit_share = sof_foyda / 2

    tikkan = {uid: 0 for uid in config.INVESTORS}
    for e in exps:
        if e["payer_id"] in tikkan:
            tikkan[e["payer_id"]] += e["amount"]
    invs = [{"id": uid, "name": config.INVESTORS[uid], "tikkan": tikkan[uid]} for uid in config.INVESTORS]
    invs.sort(key=lambda x: x["id"])

    settlement = {"amount": 0, "payer": None, "receiver": None}
    if len(invs) == 2:
        a, b = invs
        diff = a["tikkan"] - b["tikkan"]
        settlement["amount"] = abs(diff) / 2
        if diff > 0:
            settlement["payer"], settlement["receiver"] = b["name"], a["name"]
        elif diff < 0:
            settlement["payer"], settlement["receiver"] = a["name"], b["name"]

    if start:
        top_products = sorted(store.range_product_stats(start, end), key=lambda x: x["qty"], reverse=True)[:5]
    else:
        top_products = store.month_top_products(year, month)

    return {
        "year": year, "month": month, "period": period, "label": label,
        "start": start, "end": end,
        "kirim": kirim, "delivery": delivery, "cogs": cogs,
        "qadoq_soni": ms["qadoq_soni"], "qadoq_sotuv": ms["qadoq_sotuv"],
        "qadoq_tannarxi": qadoq_tannarxi,
        "qadoq_oddiy_soni": oddiy_soni, "qadoq_quti_soni": quti_soni,
        "qadoq_oddiy_tannarxi": qadoq_oddiy_tannarxi, "qadoq_quti_tannarxi": qadoq_quti_tannarxi,
        "operatsion": operatsion, "tovar_rasxod": tovar,
        "sof_foyda": sof_foyda, "profit_share": profit_share,
        "yalpi_foyda": kirim - delivery - cogs,
        "investors": invs, "settlement": settlement,
        "orders_count": ms["orders_count"], "status_filtered": ms["status_filtered"],
        "paytur": ms["paytur"], "expenses": exps,
        "cashback_used": ms.get("cashback_used", 0),
        "op_breakdown": op_breakdown,
        "top_products": top_products,
        "have_data": store.have_orders(),
    }


def compute_month(year, month):
    import calendar
    last = calendar.monthrange(year, month)[1]
    ms = store.month_sales(year, month)
    exps = db.month_expenses(year, month)
    return _compute_core(ms, exps, period=f"{year:04d}-{month:02d}",
                         label=f"{OYLAR[month]} {year}", year=year, month=month,
                         start=f"{year:04d}-{month:02d}-01", end=f"{year:04d}-{month:02d}-{last:02d}")


def compute_range(start, end, label=None):
    """Ixtiyoriy davr (YYYY-MM-DD..YYYY-MM-DD) bo'yicha hisob. Tenglashtirish auto-saqlanmaydi."""
    ms = store.range_sales(start, end)
    exps = db.range_expenses(start, end)
    return _compute_core(ms, exps, period=f"{start}_{end}",
                         label=label or f"{start} … {end}", start=start, end=end)


def auto_settlement(year, month):
    """Tenglashtirishni avtomatik tarixga yozadi (farq bo'lsa). 'to'langan' holati saqlanadi."""
    try:
        r = compute_month(year, month)
        s = r["settlement"]
        if s["amount"] > 0:
            db.save_settlement(r["period"], s["payer"], s["receiver"], s["amount"])
    except Exception:
        pass


def investor_balances():
    """Umumiy (all-time): har investorning jami tikkani + to'lanmagan tenglashtirish."""
    tikkan = {uid: 0 for uid in config.INVESTORS}
    for e in db.all_expenses():
        if e["payer_id"] in tikkan:
            tikkan[e["payer_id"]] += e["amount"]
    rows = [{"id": uid, "name": config.INVESTORS[uid], "tikkan": tikkan[uid]}
            for uid in config.INVESTORS]
    return rows, db.unpaid_settlement_balance()


# ===================== report_pdf.py =====================
"""Word-style oylik PDF (reportlab) — to'liq model B."""
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, HRFlowable)

MOCHA=colors.HexColor("#6B4226"); GOLD=colors.HexColor("#B8860B")
CREAM=colors.HexColor("#FAF4EC"); DARK=colors.HexColor("#3A2A1E")
GREEN=colors.HexColor("#3E7D5A"); RED=colors.HexColor("#B5503F")
LINE=colors.HexColor("#E5D9C8")
OYLAR=["","Yanvar","Fevral","Mart","Aprel","May","Iyun","Iyul","Avgust",
       "Sentabr","Oktabr","Noyabr","Dekabr"]


def _ss():
    ss=getSampleStyleSheet()
    ss.add(ParagraphStyle("H1",parent=ss["Title"],textColor=MOCHA,fontSize=22,spaceAfter=2,leading=26))
    ss.add(ParagraphStyle("Sub",parent=ss["Normal"],textColor=GOLD,fontSize=12,alignment=TA_CENTER,spaceAfter=12))
    ss.add(ParagraphStyle("Sec",parent=ss["Heading2"],textColor=MOCHA,fontSize=12.5,spaceBefore=12,spaceAfter=5))
    ss.add(ParagraphStyle("Foot",parent=ss["Normal"],textColor=colors.grey,fontSize=8,alignment=TA_CENTER))
    return ss


def _kv(rows, hl=None, colors_map=None):
    t=Table(rows,colWidths=[105*mm,65*mm])
    style=[("FONTSIZE",(0,0),(-1,-1),11),("TEXTCOLOR",(0,0),(0,-1),DARK),
           ("TEXTCOLOR",(1,0),(1,-1),MOCHA),("ALIGN",(1,0),(1,-1),"RIGHT"),
           ("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),
           ("LINEBELOW",(0,0),(-1,-2),0.4,LINE)]
    if colors_map:
        for idx,col in colors_map.items():
            style.append(("TEXTCOLOR",(1,idx),(1,idx),col))
    if hl is not None:
        style+=[("BACKGROUND",(0,hl),(-1,hl),CREAM),("FONTNAME",(0,hl),(-1,hl),"Helvetica-Bold"),
                ("LINEABOVE",(0,hl),(-1,hl),1,GOLD)]
    t.setStyle(TableStyle(style)); return t


def _chart(r):
    """Brend rangidagi gorizontal ustun diagramma — pul oqimi (kirim/xarajat/foyda)."""
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import HorizontalBarChart
    SAND  = colors.HexColor("#C7A98B"); SAND2 = colors.HexColor("#D8C3A5")
    labels = ["Kirim", "Tovar t/n", "Operatsion", "Qadoq t/n", "Delivery", "Sof foyda"]
    vals   = [r["kirim"], r["cogs"], r["operatsion"], r["qadoq_tannarxi"],
              r["delivery"], r["sof_foyda"]]
    bar_cols = [GOLD, MOCHA, SAND, SAND2, LINE, (GREEN if r["sof_foyda"] >= 0 else RED)]
    d = Drawing(440, 150)
    bc = HorizontalBarChart()
    bc.x = 96; bc.y = 8; bc.height = 132; bc.width = 296
    bc.data = [vals]
    bc.barWidth = 9
    bc.groupSpacing = 9
    bc.valueAxis.valueMin = min(0, min(vals))
    bc.valueAxis.visible = False
    bc.valueAxis.gridStrokeColor = colors.white
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.strokeColor = LINE
    bc.categoryAxis.labels.fontName = "Helvetica"
    bc.categoryAxis.labels.fontSize = 8.5
    bc.categoryAxis.labels.fillColor = DARK
    bc.categoryAxis.labels.dx = -3
    bc.bars.strokeColor = None
    for i, c in enumerate(bar_cols):
        bc.bars[(0, i)].fillColor = c
    d.add(bc)
    return d


def generate(r, out_path):
    ss=_ss(); doc=SimpleDocTemplate(out_path,pagesize=A4,topMargin=18*mm,bottomMargin=15*mm,
        leftMargin=20*mm,rightMargin=20*mm,title=f"BabyDiary {r['period']}")
    el=[]
    el.append(Paragraph("BabyDiary",ss["H1"]))
    _sub = r.get("label") or (f"{OYLAR[r['month']]} {r['year']}" if r.get("month") else r.get("period",""))
    el.append(Paragraph(f"Moliyaviy Hisobot — {_sub}",ss["Sub"]))
    el.append(HRFlowable(width="100%",thickness=1.5,color=GOLD,spaceAfter=8))

    # 1. Sof foyda hisobi
    el.append(Paragraph("Sof foyda hisobi",ss["Sec"]))
    sof=r["sof_foyda"]
    el.append(_kv([
        ["Kirim (delivery + qadoq ichida)", somm(r["kirim"])],
        ["− Delivery (kuryerga)", "− "+somm(r["delivery"])],
        ["− Tovar tannarxi (COGS)", "− "+somm(r["cogs"])],
        [f"− Qadoq tannarxi ({r.get('qadoq_oddiy_soni',0)} oddiy + {r.get('qadoq_quti_soni',0)} qutili)", "− "+somm(r["qadoq_tannarxi"])],
        ["− Operatsion rasxod", "− "+somm(r["operatsion"])],
        ["SOF FOYDA", somm(sof)],
    ], hl=5, colors_map={5: GREEN if sof>=0 else RED}))

    if not r.get("have_data"):
        el.append(Spacer(1,3))
        el.append(Paragraph("DIQQAT: savdo fayllari (orders.json) hali yuklanmagan — kirim 0.",ss["Foot"]))
    elif not r.get("status_filtered"):
        el.append(Spacer(1,3))
        el.append(Paragraph("Eslatma: babydiary.db yuklanmagani uchun barcha buyurtmalar hisobga olindi (status filtri yo'q).",ss["Foot"]))

    # Diagramma (brend rangida)
    try:
        el.append(Spacer(1,6))
        el.append(_chart(r))
        el.append(Spacer(1,2))
    except Exception:
        pass

    if r.get("cashback_used"):
        el.append(Paragraph(f"Eslatma: cashback ishlatilgan {somm(r['cashback_used'])} — kirimdan ayrilgan.",ss["Foot"]))

    # Operatsion rasxod — kategoriyalar kesimi
    if r.get("op_breakdown"):
        el.append(Paragraph("Operatsion rasxod (kategoriya bo'yicha)",ss["Sec"]))
        el.append(_kv([[k, somm(v)] for k, v in r["op_breakdown"].items()]))

    # 2. Foyda taqsimoti
    el.append(Paragraph("Foyda taqsimoti (50/50)",ss["Sec"]))
    el.append(_kv([[i["name"], somm(r["profit_share"])] for i in r["investors"]]))

    # 3. Kim qancha pul tikkan
    el.append(Paragraph("Kim qancha pul tikkan (bu oy)",ss["Sec"]))
    rows=[[i["name"], somm(i["tikkan"])] for i in r["investors"]]
    el.append(_kv(rows))

    # 4. Tenglashtirish
    el.append(Paragraph("Tenglashtirish",ss["Sec"]))
    s=r["settlement"]
    if s["amount"]>0:
        box=Table([[Paragraph(f"<b>{s['payer']} → {s['receiver']}</b> :  <b>{somm(s['amount'])}</b>",
            ParagraphStyle("b",textColor=DARK,fontSize=13,alignment=TA_CENTER))]],colWidths=[170*mm])
        box.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),CREAM),("BOX",(0,0),(-1,-1),1,GOLD),
            ("TOPPADDING",(0,0),(-1,-1),11),("BOTTOMPADDING",(0,0),(-1,-1),11)]))
        el.append(box)
        el.append(Spacer(1,3))
        el.append(Paragraph(f"Ko'p tikkan {s['receiver']}ga farqning yarmi qaytadi.",ss["Foot"]))
    else:
        el.append(Paragraph("Teng — tenglashtirish kerak emas.",ss["Normal"]))

    # 5. Qo'shimcha ko'rsatkichlar
    el.append(Paragraph("Qo'shimcha ko'rsatkichlar",ss["Sec"]))
    extra=[
        ["Buyurtmalar soni", f"{r['orders_count']} ta"],
        ["Yalpi foyda (kirim − delivery − COGS)", somm(r["yalpi_foyda"])],
        ["Delivery jami (kuryerga)", somm(r["delivery"])],
        ["Qadoq sotuvi", somm(r["qadoq_sotuv"])],
        ["Ombor qiymati (qotgan pul)", somm(store_inventory())],
    ]
    el.append(_kv(extra))

    # Eng ko'p sotilgan mahsulotlar
    if r.get("top_products"):
        el.append(Paragraph("Eng ko'p sotilgan mahsulotlar",ss["Sec"]))
        data=[["#","Mahsulot","Soni","Tushum","Foyda"]]
        for n,p in enumerate(r["top_products"],1):
            data.append([str(n), p["name"], f"{p['qty']} ta", somm(p["revenue"]), somm(p["foyda"])])
        t=Table(data,colWidths=[8*mm,70*mm,22*mm,35*mm,35*mm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),MOCHA),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTSIZE",(0,0),(-1,-1),8.5),("ALIGN",(2,1),(-1,-1),"RIGHT"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,CREAM]),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
        el.append(t)

    # 6. To'lov turlari
    if r["paytur"]:
        el.append(Paragraph("To'lov turlari",ss["Sec"]))
        el.append(_kv([[k, somm(v)] for k,v in r["paytur"].items()]))

    # 7. Operatsion rasxodlar ro'yxati
    op_rows=[e for e in r["expenses"] if e["kind"]=="operatsion"]
    tv_rows=[e for e in r["expenses"] if e["kind"]=="tovar"]
    if op_rows or tv_rows:
        el.append(Paragraph("Rasxodlar ro'yxati",ss["Sec"]))
        data=[["#","Tur","Summa","Izoh","Kim"]]
        for n,e in enumerate(op_rows+tv_rows,1):
            tur="Operatsion" if e["kind"]=="operatsion" else "Tovar"
            data.append([str(n),tur,somm(e["amount"]),e["note"] or "—",e["payer_name"] or "—"])
        t=Table(data,colWidths=[8*mm,24*mm,34*mm,64*mm,30*mm])
        t.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),MOCHA),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTSIZE",(0,0),(-1,-1),8.5),("ALIGN",(2,1),(2,-1),"RIGHT"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,CREAM]),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
        el.append(t)

    el.append(Spacer(1,12))
    el.append(HRFlowable(width="100%",thickness=0.5,color=LINE))
    el.append(Paragraph(f"Hurmatli Ibrohim va Rustam — biznesingizga omad!  ·  "
        f"{datetime.now(config.TZ).strftime('%Y-%m-%d %H:%M')}  ·  BabyDiary moliya boti",ss["Foot"]))
    doc.build(el); return out_path


def store_inventory():
    return store.inventory_value()


# ===================== report_xlsx.py =====================
"""Excel hisobot (openpyxl) — formula shabloni: inputlar ko'k, Sof foyda formula."""

def make_xlsx(r, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    BLUE = Font(name="Arial", size=11, color="0000FF")          # input (o'zgartirsa bo'ladi)
    BLACK = Font(name="Arial", size=11, color="000000")
    BOLD = Font(name="Arial", size=11, bold=True)
    TITLE = Font(name="Arial", size=14, bold=True, color="6B4226")
    HEADF = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    HEADFILL = PatternFill("solid", fgColor="6B4226")
    MONEY = '#,##0;(#,##0);-'
    thin = Side(style="thin", color="D9CFC0")
    border = Border(bottom=thin)

    wb = Workbook()
    # --- 1. Hisobot (P&L formula shabloni) ---
    ws = wb.active; ws.title = "Hisobot"
    ws["A1"] = "BabyDiary — Moliyaviy hisobot"; ws["A1"].font = TITLE
    ws["A2"] = r.get("label", ""); ws["A2"].font = Font(name="Arial", size=11, italic=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    rows = [
        ("Kirim (so'm)",            r["kirim"],          "input"),
        ("− Delivery (kuryerga)",   r["delivery"],       "input"),
        ("− Tovar tannarxi (COGS)", r["cogs"],           "input"),
        ("− Qadoq tannarxi",        r["qadoq_tannarxi"], "input"),
        ("− Operatsion rasxod",     r["operatsion"],     "input"),
    ]
    start_row = 4
    for i, (label, val, _) in enumerate(rows):
        rr = start_row + i
        ws[f"A{rr}"] = label; ws[f"A{rr}"].font = BLACK
        c = ws[f"B{rr}"]; c.value = int(val); c.font = BLUE; c.number_format = MONEY
    sof_row = start_row + len(rows)
    ws[f"A{sof_row}"] = "SOF FOYDA"; ws[f"A{sof_row}"].font = BOLD
    sof = ws[f"B{sof_row}"]
    sof.value = f"=B{start_row}-B{start_row+1}-B{start_row+2}-B{start_row+3}-B{start_row+4}"
    sof.font = BOLD; sof.number_format = MONEY
    for col in ("A", "B"):
        ws[f"{col}{sof_row}"].border = Border(top=Side(style="medium", color="B8860B"))
    sh_row = sof_row + 1
    ws[f"A{sh_row}"] = "Har investorga (50/50)"; ws[f"A{sh_row}"].font = BLACK
    ws[f"B{sh_row}"] = f"=B{sof_row}/2"; ws[f"B{sh_row}"].font = BLACK; ws[f"B{sh_row}"].number_format = MONEY
    ob_row = sh_row + 2
    ws[f"A{ob_row}"] = "Buyurtmalar soni"; ws[f"A{ob_row}"].font = BLACK
    ws[f"B{ob_row}"] = r["orders_count"]; ws[f"B{ob_row}"].font = BLACK
    ws[f"A{ob_row+1}"] = "(ko'k raqamlarni o'zgartirsangiz, SOF FOYDA avtomatik qayta hisoblanadi)"
    ws[f"A{ob_row+1}"].font = Font(name="Arial", size=9, italic=True, color="808080")

    # --- 2. Tovarlar (eng ko'p sotilgan) ---
    ws2 = wb.create_sheet("Tovarlar")
    heads = ["Mahsulot", "Sotilgan (dona)", "Savdo (so'm)", "Foyda (so'm)"]
    for j, h in enumerate(heads, 1):
        c = ws2.cell(row=1, column=j, value=h); c.font = HEADF; c.fill = HEADFILL
    widths2 = [34, 16, 16, 16]
    for j, w in enumerate(widths2, 1):
        ws2.column_dimensions[chr(64+j)].width = w
    for i, p in enumerate(r.get("top_products", []), start=2):
        ws2.cell(row=i, column=1, value=p["name"]).font = BLACK
        ws2.cell(row=i, column=2, value=p["qty"]).font = BLACK
        c3 = ws2.cell(row=i, column=3, value=p["revenue"]); c3.font = BLACK; c3.number_format = MONEY
        c4 = ws2.cell(row=i, column=4, value=p["foyda"]); c4.font = BLACK; c4.number_format = MONEY

    # --- 3. Rasxodlar ---
    ws3 = wb.create_sheet("Rasxodlar")
    heads3 = ["Sana", "Tur", "Kategoriya", "Summa (so'm)", "Izoh"]
    for j, h in enumerate(heads3, 1):
        c = ws3.cell(row=1, column=j, value=h); c.font = HEADF; c.fill = HEADFILL
    for j, w in enumerate([18, 12, 14, 16, 30], 1):
        ws3.column_dimensions[chr(64+j)].width = w
    for i, e in enumerate(r.get("expenses", []), start=2):
        ws3.cell(row=i, column=1, value=(e.get("created_at") or "")[:10]).font = BLACK
        ws3.cell(row=i, column=2, value=("Tovar" if e.get("kind")=="tovar" else "Operatsion")).font = BLACK
        ws3.cell(row=i, column=3, value=e.get("subcat") or "—").font = BLACK
        c4 = ws3.cell(row=i, column=4, value=int(e.get("amount",0))); c4.font = BLACK; c4.number_format = MONEY
        ws3.cell(row=i, column=5, value=e.get("note") or "").font = BLACK

    wb.save(out_path)
    return out_path


# ===================== bot.py =====================
"""Moliya boti — handlerlar (python-telegram-bot v21)."""
import os, tempfile
from datetime import datetime
from functools import wraps
from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import ContextTypes, CommandHandler, MessageHandler, CallbackQueryHandler, filters

MENU = ReplyKeyboardMarkup([
    ["➕ Rasxod", "📊 Holat"],
    ["📄 Hisobot (PDF)", "👤 Balans"],
    ["📦 Ombor", "📈 Tahlil"],
    ["✏️ O'zgartirish", "⚙️ Sozlamalar"],
], resize_keyboard=True)

# Bo'limga kirilganda pastki menyu shu bitta tugmaga yopiladi
BACK_MENU = ReplyKeyboardMarkup([["🏠 Asosiy menu"]], resize_keyboard=True)


def filter_menu():
    """Holat paneli: investorlar bo'yicha filtr + Asosiy menu."""
    names = list(config.INVESTORS.values())
    rows = [[f"👤 {n}" for n in names]]
    rows.append(["👥 Hammasi"])
    rows.append(["🏠 Asosiy menu"])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)

# Oqim ichida bosh menuga qaytish uchun inline tugma
HOME_BTN = InlineKeyboardButton("🏠 Asosiy menu", callback_data="home")

# Operatsion rasxod kichik kategoriyalari
OP_SUBCATS = ["Ijara", "Ish haqi", "Reklama", "Poligrafiya", "Oformleniye", "Kommunal", "Transport", "Boshqa"]


def get_op_subcats():
    """Standart + admin botdan qo'shgan operatsion kategoriyalar."""
    import json as _json
    extra = []
    try:
        raw = db.get_setting("op_subcats_extra", "")
        if raw:
            extra = [x for x in _json.loads(raw) if x]
    except Exception:
        extra = []
    out = []
    for c in OP_SUBCATS + extra:
        if c not in out:
            out.append(c)
    return out


def add_op_subcat(name):
    """Yangi operatsion kategoriya qo'shadi (botdan)."""
    import json as _json
    name = (name or "").strip()
    if not name:
        return
    try:
        raw = db.get_setting("op_subcats_extra", "")
        extra = _json.loads(raw) if raw else []
    except Exception:
        extra = []
    if name not in OP_SUBCATS and name not in extra:
        extra.append(name)
        db.set_setting("op_subcats_extra", _json.dumps(extra, ensure_ascii=False))


def admin_only(func):
    @wraps(func)
    async def w(update, context):
        u = update.effective_user
        if not u or u.id not in config.ADMIN_IDS:
            if update.message:
                await update.message.reply_text("⛔ Bu bot faqat investorlar uchun.")
            return
        return await func(update, context)
    return w


@admin_only
async def go_home(update, context):
    """Istalgan oqimni bekor qilib, asosiy menuga qaytaradi."""
    context.user_data.clear()
    await update.message.reply_text("🏠 Asosiy menu", reply_markup=MENU)


@admin_only
async def home_cb(update, context):
    """Inline '🏠 Bosh menu' tugmasi — oqimni bekor qilib menuga qaytaradi."""
    q = update.callback_query
    context.user_data.clear()
    await q.answer("🏠")
    try:
        await q.edit_message_text("🏠 Asosiy menuga qaytdingiz.")
    except Exception:
        pass
    await context.bot.send_message(q.message.chat_id, "Asosiy menu 👇", reply_markup=MENU)


def parse_amount(t):
    t = (t or "").lower().replace(" ", "").replace("so'm","").replace("som","")
    mult = 1
    for suf in ("mln","млн"):
        if t.endswith(suf): t=t[:-len(suf)]; mult=1_000_000; break
    else:
        if t.endswith("k") or t.endswith("ming"):
            t=t.rstrip("k").replace("ming",""); mult=1_000
    t=t.replace(",",".")
    try: return int(round(float(t)*mult))
    except ValueError: return None


async def notify_partner(context, actor_id, text, photo=None):
    """Rasxod qo'shilsa, ikkinchi investorga xabar (chek rasmi bilan ham)."""
    for uid in config.ADMIN_IDS:
        if uid != actor_id:
            try:
                if photo:
                    await context.bot.send_photo(uid, photo, caption=text)
                else:
                    await context.bot.send_message(uid, text)
            except Exception:
                pass


# ---------- /start ----------
@admin_only
async def start(update, context):
    await update.message.reply_text("BabyDiary moliya boti 💼", reply_markup=MENU)


# ---------- Rasxod oqimi ----------
@admin_only
async def rasxod_start(update, context):
    context.user_data["flow"] = "rasxod_summa"
    kb = InlineKeyboardMarkup([[HOME_BTN]])
    await update.message.reply_text("➕ Rasxod summasini yozing (masalan: 500000 yoki 1.5mln):", reply_markup=kb)


async def rasxod_text(update, context):
    flow = context.user_data.get("flow")
    txt = (update.message.text or "").strip()
    if flow == "rasxod_summa":
        amt = parse_amount(txt)
        if not amt or amt <= 0:
            await update.message.reply_text("❌ Summa noto'g'ri. Qaytadan yozing:",
                reply_markup=InlineKeyboardMarkup([[HOME_BTN]]))
            return
        context.user_data["r_amount"] = amt
        context.user_data["flow"] = "rasxod_izoh"
        await update.message.reply_text(f"Summa: {somm(amt)}\n\nEndi izoh yozing (masalan: ijara, tovar partiya):",
            reply_markup=InlineKeyboardMarkup([[HOME_BTN]]))
    elif flow == "rasxod_izoh":
        context.user_data["r_note"] = txt or "—"
        context.user_data["flow"] = "rasxod_rasm"
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➡️ Rasmsiz davom etish", callback_data="rk_norasm")],
            [HOME_BTN],
        ])
        await update.message.reply_text(
            "📷 Chek/rasm bo'lsa yuboring, yoki rasmsiz davom eting:", reply_markup=kb)
    elif flow == "rasxod_rasm":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("➡️ Rasmsiz davom etish", callback_data="rk_norasm")],
            [HOME_BTN],
        ])
        await update.message.reply_text(
            "📷 Chek rasmini yuboring yoki rasmsiz davom eting:", reply_markup=kb)
    elif flow == "rasxod_newsubcat":
        name = txt.strip()
        if not name:
            await update.message.reply_text("❌ Nomi bo'sh. Qaytadan yozing:")
            return
        add_op_subcat(name)
        context.user_data["r_subcat"] = name
        await _rasxod_confirm(context, update.effective_chat.id)


async def _ask_rasxod_kind(context, chat_id):
    """Rasxod turini so'rash (tovar/operatsion) — rasm bosqichidan keyin chaqiriladi."""
    context.user_data["flow"] = "rasxod_tur"
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🛍 Tovar", callback_data="rk_tovar"),
         InlineKeyboardButton("🏠 Operatsion", callback_data="rk_operatsion")],
        [HOME_BTN],
    ])
    await context.bot.send_message(chat_id, "Rasxod turi?", reply_markup=kb)


@admin_only
async def rasxod_photo(update, context):
    """Rasxod oqimida yuborilgan chek rasmini qabul qiladi."""
    if context.user_data.get("flow") != "rasxod_rasm":
        return
    context.user_data["r_photo"] = update.message.photo[-1].file_id
    await update.message.reply_text("📷 Chek rasmi qabul qilindi.")
    await _ask_rasxod_kind(context, update.effective_chat.id)


@admin_only
async def rasxod_norasm_cb(update, context):
    q = update.callback_query
    if context.user_data.get("flow") != "rasxod_rasm":
        await q.answer(); return
    context.user_data["r_photo"] = None
    await q.answer()
    try:
        await q.edit_message_text("📷 Rasmsiz davom etilyapti.")
    except Exception:
        pass
    await _ask_rasxod_kind(context, q.message.chat_id)


@admin_only
async def rasxod_kind_cb(update, context):
    q = update.callback_query
    if context.user_data.get("flow") != "rasxod_tur":
        await q.answer(); return
    kind = "tovar" if q.data == "rk_tovar" else "operatsion"
    context.user_data["r_kind"] = kind
    await q.answer()
    if kind == "operatsion":
        # Operatsion bo'lsa — kichik kategoriya so'raymiz
        context.user_data["flow"] = "rasxod_subcat"
        subcats = get_op_subcats()
        context.user_data["r_subcats"] = subcats
        rows = []
        row = []
        for i, sc in enumerate(subcats):
            row.append(InlineKeyboardButton(sc, callback_data=f"rsub_{i}"))
            if len(row) == 2:
                rows.append(row); row = []
        if row: rows.append(row)
        rows.append([InlineKeyboardButton("➕ Yangi kategoriya", callback_data="rsub_new")])
        rows.append([HOME_BTN])
        try:
            await q.edit_message_text("🏠 Operatsion — kategoriyani tanlang:",
                reply_markup=InlineKeyboardMarkup(rows))
        except Exception:
            await context.bot.send_message(q.message.chat_id, "🏠 Operatsion — kategoriyani tanlang:",
                reply_markup=InlineKeyboardMarkup(rows))
    else:
        context.user_data["r_subcat"] = ""
        await _rasxod_confirm(context, q.message.chat_id, edit_q=q)


@admin_only
async def rasxod_subcat_cb(update, context):
    q = update.callback_query
    if context.user_data.get("flow") != "rasxod_subcat":
        await q.answer(); return
    if q.data == "rsub_new":
        context.user_data["flow"] = "rasxod_newsubcat"
        await q.answer()
        try:
            await q.edit_message_text("✍️ Yangi kategoriya nomini yozing:")
        except Exception:
            await context.bot.send_message(q.message.chat_id, "✍️ Yangi kategoriya nomini yozing:")
        return
    try:
        idx = int(q.data.replace("rsub_", ""))
        subcats = context.user_data.get("r_subcats") or get_op_subcats()
        context.user_data["r_subcat"] = subcats[idx]
    except Exception:
        context.user_data["r_subcat"] = "Boshqa"
    await q.answer()
    await _rasxod_confirm(context, q.message.chat_id, edit_q=q)


async def _rasxod_confirm(context, chat_id, edit_q=None):
    """Saqlashdan oldin preview + tasdiq tugmasi."""
    context.user_data["flow"] = "rasxod_confirm"
    amt = context.user_data.get("r_amount")
    note = context.user_data.get("r_note", "—")
    kind = context.user_data.get("r_kind")
    subcat = context.user_data.get("r_subcat", "")
    photo = context.user_data.get("r_photo")
    tur_txt = "🛍 Tovar" if kind == "tovar" else f"🏠 Operatsion · {subcat}"
    rasm_txt = "📷 bor" if photo else "📷 yo'q"
    text = (f"📝 *Tekshiring:*\n\n"
            f"💰 Summa: {somm(amt)}\n"
            f"📂 Tur: {tur_txt}\n"
            f"📄 Izoh: {md_escape(note)}\n"
            f"🧾 Chek: {rasm_txt}\n\n"
            f"Saqlaymizmi?")
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ Saqlash", callback_data="rsave"),
         InlineKeyboardButton("❌ Bekor", callback_data="home")],
    ])
    if edit_q is not None:
        try:
            await edit_q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
            return
        except Exception:
            pass
    await context.bot.send_message(chat_id, text, parse_mode="Markdown", reply_markup=kb)


@admin_only
async def rasxod_save_cb(update, context):
    q = update.callback_query
    if context.user_data.get("flow") != "rasxod_confirm":
        await q.answer(); return
    amt = context.user_data.get("r_amount"); note = context.user_data.get("r_note","—")
    kind = context.user_data.get("r_kind"); subcat = context.user_data.get("r_subcat","")
    photo = context.user_data.get("r_photo")
    uid = q.from_user.id; name = config.INVESTORS.get(uid,"")
    rid = db.add_expense(amt, note, kind, uid, name, photo, subcat)
    context.user_data.clear()
    tur_txt = "🛍 Tovar" if kind=="tovar" else f"🏠 Operatsion · {subcat}"
    rasm_txt = "\n📷 Chek biriktirildi" if photo else ""
    await q.answer("✅ Saqlandi")
    await q.edit_message_text(f"✅ Rasxod #{rid} saqlandi:\n{somm(amt)} — {note}\nTur: {tur_txt}\nKim: {name}{rasm_txt}")
    await context.bot.send_message(q.message.chat_id, "Tayyor.", reply_markup=MENU)
    await notify_partner(context, uid, f"🔔 {name} rasxod qo'shdi:\n{somm(amt)} — {note} ({tur_txt})", photo=photo)


# ---------- Davr (oy) tugmalari — faqat ma'lumot bor oylar ----------
def data_months():
    """Rasxod yoki buyurtma bor oylar: [(year,month), ...] (eskidan yangiga)."""
    months = set()
    try:
        for e in db.all_expenses():
            ym = (e.get("created_at") or "")[:7]
            if len(ym) == 7: months.add(ym)
    except Exception:
        pass
    try:
        for o in store.load_orders():
            ym = (o.get("date") or "")[:7]
            if len(ym) == 7: months.add(ym)
    except Exception:
        pass
    now = datetime.now(config.TZ)
    months.add(f"{now.year:04d}-{now.month:02d}")   # joriy oy doim bo'lsin
    out = []
    for ym in sorted(months):
        try:
            y, m = ym.split("-"); out.append((int(y), int(m)))
        except Exception:
            pass
    return out

def months_kb(prefix, home=True):
    """Faqat ma'lumot bor oylar — oy nomi bilan tugma (eng yangi tepada)."""
    rows = [[InlineKeyboardButton(f"📅 {OYLAR[m]} {y}", callback_data=f"{prefix}_{y}_{m}")]
            for (y, m) in reversed(data_months())]
    if home:
        rows.append([HOME_BTN])
    return InlineKeyboardMarkup(rows)


# ---------- Holat (jonli) ----------
def holat_text(year, month):
    r = finance.compute_month(year, month)
    sof = r["sof_foyda"]
    L = [f"📊 *{OYLAR[r['month']]} {r['year']}* — holat", ""]
    if not r["have_data"]:
        L.append("⚠️ Savdo fayllari yo'q (kirim 0).")
        L.append("")
    L += [
        f"Kirim:  {somm(r['kirim'])}",
        f"− Delivery: {somm(r['delivery'])} (kuryerga)",
        f"− Tovar tannarxi: {somm(r['cogs'])}",
        f"− Qadoq tannarxi: {somm(r['qadoq_tannarxi'])}",
        f"− Operatsion: {somm(r['operatsion'])}",
        f"*SOF FOYDA: {somm(sof)}*",
        f"Har biriga: {somm(r['profit_share'])}",
        "",
        f"🛍 Tovarga kiritilgan pul: {somm(r.get('tovar_rasxod', 0))}",
    ]
    if r.get("cashback_used"):
        L.append(f"_(cashback ishlatilgan: {somm(r['cashback_used'])} — kirimdan ayrilgan)_")

    # Oldingi oy bilan taqqoslash (sof foyda)
    py, pm = (year - 1, 12) if month == 1 else (year, month - 1)
    pr = finance.compute_month(py, pm)
    if pr.get("have_data") and pr["sof_foyda"] != 0:
        diff = sof - pr["sof_foyda"]
        pct = diff / abs(pr["sof_foyda"]) * 100
        arrow = "📈" if diff >= 0 else "📉"
        sign = "+" if diff >= 0 else "−"
        L.append(f"{arrow} O'tgan oyga nisbatan: {sign}{somm(abs(diff))} ({pct:+.0f}%)")

    # Mijozlarda turgan cashback — MAJBURIYAT, foydadan AYRILMAYDI (faqat ma'lumot)
    ocb = store.outstanding_cashback()
    if ocb:
        L.append(f"_ℹ️ Mijozlarda turgan cashback: {somm(ocb)} (foydadan ayrilmaydi, ishlatilganda hisobga olinadi)_")

    L += [
        "",
        "Kim tikkan (bu oy):",
    ]
    for i in r["investors"]:
        L.append(f"  • {i['name']}: {somm(i['tikkan'])}")
    return "\n".join(L)


def _expense_list_content(year, month, editable=False, filter_uid=None):
    """Rasxod ro'yxati matni + klaviaturasi. filter_uid bo'lsa faqat o'sha investorniki."""
    exps = db.month_expenses(year, month)
    if filter_uid:
        exps = [e for e in exps if e.get("payer_id") == filter_uid]
    who = config.INVESTORS.get(filter_uid, "") if filter_uid else ""
    if editable:
        title = "o'zgartirish (rasxod tanlang)"
    elif filter_uid:
        title = f"{who} rasxodlari"
    else:
        title = "rasxodlar ro'yxati"
    if not exps:
        L = [f"📋 *{OYLAR[month]} {year}* — {title}:\n", "_Rasxod yo'q._"]
    else:
        L = [f"📋 *{OYLAR[month]} {year}* — {title}:\n"]
    kb = []
    total = 0
    for e in exps:
        total += e["amount"]
        tur = "🛍 Tovar" if e["kind"] == "tovar" else f"🏠 {e.get('subcat') or 'Operatsion'}"
        cam = " 📷" if e.get("photo_id") else ""
        pend = " ⏳" if (e.get("del_req_by") or e.get("edit_req_by")) else ""
        L.append(f"#{e['id']} {md_escape(tur)}{cam}{pend}\n"
                 f"   💰 {somm(e['amount'])}  ·  👤 {md_escape(e['payer_name'])}\n"
                 f"   📄 {md_escape(e['note'])}")
        if editable:
            if e.get("del_req_by") or e.get("edit_req_by"):
                kb.append([InlineKeyboardButton(f"⏳ #{e['id']} — tasdiq kutilmoqda", callback_data=f"delreq_{e['id']}")])
            else:
                kb.append([
                    InlineKeyboardButton(f"✏️ O'zgartirish #{e['id']}", callback_data=f"editreq_{e['id']}"),
                    InlineKeyboardButton(f"🗑 #{e['id']}", callback_data=f"delreq_{e['id']}"),
                ])
    if exps:
        L.append(f"\n*Jami: {somm(total)}*")
    if editable:
        L.append("\n_tahrirlash / o'chirish — ikkala investor tasdig'i kerak._")
    else:
        # Oy navigatsiyasi (◀️ oldingi / keyingi ▶️) — o'tgan oylarni ko'rish uchun
        who_key = filter_uid if filter_uid else "all"
        py, pm = _shift_month(year, month, -1)
        ny, nm = _shift_month(year, month, 1)
        kb.append([
            InlineKeyboardButton(f"◀️ {OYLAR[pm]}", callback_data=f"exm_{who_key}_{py}_{pm}"),
            InlineKeyboardButton(f"{OYLAR[nm]} ▶️", callback_data=f"exm_{who_key}_{ny}_{nm}"),
        ])
    markup = InlineKeyboardMarkup(kb) if kb else None
    return "\n".join(L), markup


async def _send_expense_list(context, chat_id, year, month, editable=False, filter_uid=None):
    text, markup = _expense_list_content(year, month, editable, filter_uid)
    await context.bot.send_message(chat_id, text, parse_mode="Markdown", reply_markup=markup)


async def send_filtered_expenses(update, context, filter_uid):
    """Panel tugmasi bosilganda — o'sha investor rasxodlari (joriy oy). (Yordamchi, dekoratorsiz.)"""
    u = update.effective_user
    if not u or u.id not in config.ADMIN_IDS:
        return
    now = datetime.now(config.TZ)
    text, markup = _expense_list_content(now.year, now.month, editable=False, filter_uid=filter_uid)
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=markup)


@admin_only
async def expense_month_cb(update, context):
    """Filtrlangan rasxod ro'yxatida oy almashtirish."""
    q = update.callback_query
    parts = q.data.split("_")           # exm_{who}_{y}_{m}
    who = parts[1]; y = int(parts[2]); m = int(parts[3])
    filter_uid = None if who == "all" else int(who)
    await q.answer()
    text, markup = _expense_list_content(y, m, editable=False, filter_uid=filter_uid)
    try:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=markup)
    except Exception:
        pass


@admin_only
async def ozgartirish(update, context):
    """O'zgartirish — tahrirlash/o'chirish tugmalari bilan rasxodlar ro'yxati."""
    now = datetime.now(config.TZ)
    await _send_expense_list(context, update.effective_chat.id, now.year, now.month, editable=True)


@admin_only
async def holat(update, context):
    now = datetime.now(config.TZ)
    finance.auto_settlement(now.year, now.month)
    await update.message.reply_text(
        holat_text(now.year, now.month), parse_mode="Markdown",
        reply_markup=months_kb("hol"))


@admin_only
async def holat_nav_cb(update, context):
    q = update.callback_query
    _, y, m = q.data.split("_"); y, m = int(y), int(m)
    finance.auto_settlement(y, m)
    await q.answer()
    try:
        await q.edit_message_text(holat_text(y, m), parse_mode="Markdown",
            reply_markup=months_kb("hol"))
    except Exception:
        pass


# ---------- Hisobot — davr tanlash (Bugungi / 7 kun / Oylik / Sana oralig'i) ----------
def _period_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 Bugungi", callback_data="per_today"),
         InlineKeyboardButton("🗓 So'nggi 7 kun", callback_data="per_7")],
        [InlineKeyboardButton("📆 Oylik", callback_data="per_month")],
        [InlineKeyboardButton("✍️ Sana oralig'i", callback_data="per_custom")],
        [HOME_BTN],
    ])


@admin_only
async def hisobot(update, context):
    await update.message.reply_text(
        "📄 *Hisobot* — qaysi davr?", parse_mode="Markdown", reply_markup=_period_kb())


def _range_label(start, end):
    import calendar
    try:
        sy, sm, sd = map(int, start.split("-"))
        ey, em, ed = map(int, end.split("-"))
        if start == end:
            return start
        if sy == ey and sm == em and sd == 1 and ed == calendar.monthrange(sy, sm)[1]:
            return f"{OYLAR[sm]} {sy}"
    except Exception:
        pass
    return f"{start} … {end}"


def range_summary_text(r):
    L = [f"📄 *Hisobot — {md_escape(r['label'])}*", ""]
    if not r.get("orders_count"):
        L.append("_Bu davrda buyurtma yo'q._")
    L += [
        f"Kirim: {somm(r['kirim'])}",
        f"− Delivery: {somm(r['delivery'])}",
        f"− Tovar tannarxi: {somm(r['cogs'])}",
        f"− Qadoq tannarxi: {somm(r['qadoq_tannarxi'])}",
        f"− Operatsion: {somm(r['operatsion'])}",
        f"*SOF FOYDA: {somm(r['sof_foyda'])}*",
        "",
        f"🛍 Tovarga kiritilgan pul: {somm(r.get('tovar_rasxod', 0))}",
        f"Buyurtmalar: {r['orders_count']} ta",
    ]
    return "\n".join(L)


def _dl_kb(start, end):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📄 PDF", callback_data=f"rgpdf_{start}_{end}"),
         InlineKeyboardButton("📊 Excel", callback_data=f"rgxls_{start}_{end}")],
        [InlineKeyboardButton("◀️ Davrlar", callback_data="per_back"), HOME_BTN],
    ])


async def _show_range(q_or_msg, start, end, edit=False):
    r = finance.compute_range(start, end, _range_label(start, end))
    txt = range_summary_text(r)
    kb = _dl_kb(start, end)
    if edit:
        await q_or_msg.edit_message_text(txt, parse_mode="Markdown", reply_markup=kb)
    else:
        await q_or_msg.reply_text(txt, parse_mode="Markdown", reply_markup=kb)


@admin_only
async def per_today_cb(update, context):
    q = update.callback_query; await q.answer()
    t = datetime.now(config.TZ).date().isoformat()
    await _show_range(q, t, t, edit=True)


@admin_only
async def per_7_cb(update, context):
    from datetime import timedelta
    q = update.callback_query; await q.answer()
    end = datetime.now(config.TZ).date()
    start = end - timedelta(days=6)
    await _show_range(q, start.isoformat(), end.isoformat(), edit=True)


@admin_only
async def per_month_cb(update, context):
    q = update.callback_query; await q.answer()
    await q.edit_message_text("📆 *Oylik* — qaysi oy?\n\nFaqat ma'lumot bor oylar 👇",
        parse_mode="Markdown", reply_markup=months_kb("repdl"))


@admin_only
async def per_back_cb(update, context):
    q = update.callback_query; await q.answer()
    await q.edit_message_text("📄 *Hisobot* — qaysi davr?", parse_mode="Markdown", reply_markup=_period_kb())


def _shift_month(y, m, delta):
    m2 = m + delta
    y2 = y + (m2 - 1) // 12
    m2 = (m2 - 1) % 12 + 1
    return y2, m2


def _cal_kb(mode, year, month, start=""):
    """Inline kalendar. mode: 's' (boshlanish) yoki 'e' (tugash)."""
    import calendar as _cal
    rows = []
    ny1, nm1 = _shift_month(year, month, -1)
    ny2, nm2 = _shift_month(year, month, 1)
    rows.append([
        InlineKeyboardButton("◀️", callback_data=f"caln_{mode}_{ny1}-{nm1}_{start}"),
        InlineKeyboardButton(f"{OYLAR[month]} {year}", callback_data="cal_nop"),
        InlineKeyboardButton("▶️", callback_data=f"caln_{mode}_{ny2}-{nm2}_{start}"),
    ])
    rows.append([InlineKeyboardButton(d, callback_data="cal_nop")
                 for d in ["Du", "Se", "Ch", "Pa", "Ju", "Sh", "Ya"]])
    for week in _cal.Calendar(firstweekday=0).monthdayscalendar(year, month):
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="cal_nop"))
            else:
                ds = f"{year:04d}-{month:02d}-{day:02d}"
                label = f"·{day}·" if (mode == "e" and start == ds) else str(day)
                row.append(InlineKeyboardButton(label, callback_data=f"cald_{mode}_{ds}_{start}"))
        rows.append(row)
    rows.append([HOME_BTN])
    return InlineKeyboardMarkup(rows)


@admin_only
async def per_custom_cb(update, context):
    q = update.callback_query; await q.answer()
    today = datetime.now(config.TZ).date()
    await q.edit_message_text(
        "📅 *Boshlanish kunini* tanlang:", parse_mode="Markdown",
        reply_markup=_cal_kb("s", today.year, today.month, ""))


@admin_only
async def cal_nop_cb(update, context):
    await update.callback_query.answer()


@admin_only
async def cal_nav_cb(update, context):
    q = update.callback_query; await q.answer()
    parts = q.data.split("_")           # caln_{mode}_{y-m}_{start}
    mode = parts[1]; ym = parts[2]; start = parts[3] if len(parts) > 3 else ""
    y, m = map(int, ym.split("-"))
    prompt = "📅 *Boshlanish kunini* tanlang:" if mode == "s" else "📅 *Tugash kunini* tanlang:"
    await q.edit_message_text(prompt, parse_mode="Markdown", reply_markup=_cal_kb(mode, y, m, start))


@admin_only
async def cal_day_cb(update, context):
    q = update.callback_query
    parts = q.data.split("_")           # cald_{mode}_{ds}_{start}
    mode = parts[1]; ds = parts[2]; start = parts[3] if len(parts) > 3 else ""
    if mode == "s":
        await q.answer()
        y, m = int(ds[:4]), int(ds[5:7])
        await q.edit_message_text(
            f"📅 Boshlanish: *{ds}*\n\nEndi *tugash kunini* tanlang:", parse_mode="Markdown",
            reply_markup=_cal_kb("e", y, m, ds))
    else:
        await q.answer()
        a, b = start, ds
        if b < a:
            a, b = b, a
        await _show_range(q, a, b, edit=True)


async def rep_custom_text(update, context):
    import re
    txt = (update.message.text or "").strip()
    dates = re.findall(r"\d{4}-\d{2}-\d{2}", txt)
    if not dates:
        await update.message.reply_text("❌ Sana formati: YYYY-MM-DD (masalan 2026-06-01 2026-06-15)")
        return
    start = dates[0]
    end = dates[1] if len(dates) > 1 else dates[0]
    if end < start:
        start, end = end, start
    context.user_data.pop("flow", None)
    await _show_range(update.message, start, end, edit=False)


@admin_only
async def report_dl_cb(update, context):
    q = update.callback_query
    _, y, m = q.data.split("_"); y, m = int(y), int(m)
    await q.answer()
    finance.auto_settlement(y, m)
    import calendar
    last = calendar.monthrange(y, m)[1]
    await _show_range(q, f"{y:04d}-{m:02d}-01", f"{y:04d}-{m:02d}-{last:02d}", edit=True)


@admin_only
async def rg_pdf_cb(update, context):
    q = update.callback_query
    parts = q.data.split("_"); start, end = parts[1], parts[2]
    await q.answer("📄 Tayyorlanmoqda...")
    r = finance.compute_range(start, end, _range_label(start, end))
    fn = f"BabyDiary_{r['period']}.pdf"
    path = os.path.join(tempfile.gettempdir(), fn)
    report_pdf.generate(r, path)
    with open(path, "rb") as f:
        await context.bot.send_document(q.message.chat_id, f, filename=fn, caption=f"📄 {r['label']}")


def _make_csv_fallback(r, path):
    """openpyxl bo'lmaganda — CSV hisobot (Excel/Google Sheets ochadi)."""
    import csv
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["BabyDiary — Moliyaviy hisobot", r.get("label", "")])
        w.writerow([])
        w.writerow(["Ko'rsatkich", "Summa (so'm)"])
        w.writerow(["Kirim", r.get("kirim", 0)])
        w.writerow(["Delivery", r.get("delivery", 0)])
        w.writerow(["Tovar tannarxi (COGS)", r.get("cogs", 0)])
        w.writerow(["Qadoq tannarxi", r.get("qadoq_tannarxi", 0)])
        w.writerow(["Operatsion rasxod", r.get("operatsion", 0)])
        w.writerow(["SOF FOYDA", r.get("sof_foyda", 0)])
        w.writerow(["Har biriga", r.get("profit_share", 0)])
        w.writerow(["Tovarga kiritilgan pul", r.get("tovar_rasxod", 0)])
        w.writerow(["Buyurtmalar soni", r.get("orders_count", 0)])
        w.writerow([])
        w.writerow(["RASXODLAR"])
        w.writerow(["#", "Tur", "Summa", "Izoh", "Kim", "Kategoriya"])
        for e in r.get("expenses", []):
            tur = "Tovar" if e.get("kind") == "tovar" else "Operatsion"
            w.writerow([e.get("id", ""), tur, e.get("amount", 0), e.get("note", ""),
                        e.get("payer_name", ""), e.get("subcat", "")])
        if r.get("top_products"):
            w.writerow([])
            w.writerow(["ENG KO'P SOTILGAN MAHSULOTLAR"])
            w.writerow(["#", "Mahsulot", "Soni", "Tushum", "Foyda"])
            for n, pr in enumerate(r["top_products"], 1):
                w.writerow([n, pr.get("name", ""), pr.get("qty", 0),
                            pr.get("revenue", 0), pr.get("foyda", 0)])


@admin_only
async def rg_xls_cb(update, context):
    q = update.callback_query
    parts = q.data.split("_"); start, end = parts[1], parts[2]
    await q.answer("📊 Tayyorlanmoqda...")
    r = finance.compute_range(start, end, _range_label(start, end))
    fn = f"BabyDiary_{r['period']}.xlsx"
    path = os.path.join(tempfile.gettempdir(), fn)
    try:
        report_xlsx.make_xlsx(r, path)
    except ImportError:
        # openpyxl o'rnatilmagan — CSV bilan (uni ham Excel ochadi)
        cpath = path.replace(".xlsx", ".csv")
        cfn = fn.replace(".xlsx", ".csv")
        _make_csv_fallback(r, cpath)
        with open(cpath, "rb") as f:
            await context.bot.send_document(
                q.message.chat_id, f, filename=cfn,
                caption=(f"📊 {r['label']} (CSV — Excel/Sheets ochadi)\n\n"
                         "To'liq (formatli) Excel uchun requirements.txt ga `openpyxl` qo'shing."),
                parse_mode="Markdown")
        return
    except Exception as e:
        await context.bot.send_message(q.message.chat_id, f"❌ Excel tayyorlashda xato: {e}")
        return
    with open(path, "rb") as f:
        await context.bot.send_document(q.message.chat_id, f, filename=fn, caption=f"📊 {r['label']}")


# ---------- Balans ----------
@admin_only
async def balans(update, context):
    rows, unpaid = finance.investor_balances()
    L = ["👤 *Balans*", ""]
    if unpaid:
        for (payer, receiver), amt in unpaid.items():
            L.append(f"🤝 {payer} → {receiver}:  *{somm(amt)}*")
    else:
        L.append("✅ Hammasi teng — hech kim hech kimga qarzdor emas.")
    await update.message.reply_text("\n".join(L), parse_mode="Markdown", reply_markup=BACK_MENU)


# ---------- Ombor ----------
@admin_only
async def ombor(update, context):
    val = store.inventory_value()
    n = len(store.load_products())
    await update.message.reply_text(
        f"📦 *Ombor qiymati*\n\nQotgan pul (tan narx × qoldiq): {somm(val)}\nMahsulot turlari: {n} ta\n\n"
        "(products.json forward qilinsa yangilanadi)", parse_mode="Markdown", reply_markup=BACK_MENU)


# ---------- Ro'yxat + o'chirish ----------
@admin_only
async def royxat(update, context):
    now = datetime.now(config.TZ)
    exps = db.month_expenses(now.year, now.month)
    if not exps:
        await update.message.reply_text("📋 Bu oyda rasxod yo'q.", reply_markup=BACK_MENU); return
    L = [f"📋 *{OYLAR[now.month]} {now.year}* rasxodlari:\n"]
    kb = []
    for e in exps:
        tur = "🛍" if e["kind"]=="tovar" else "🏠"
        sc = f" ({md_escape(e.get('subcat'))})" if e.get("subcat") else ""
        cam = " 📷" if e.get("photo_id") else ""
        pend = " ⏳" if (e.get("del_req_by") or e.get("edit_req_by")) else ""
        L.append(f"#{e['id']} {tur}{sc}{cam}{pend} {somm(e['amount'])} — {md_escape(e['note'])} ({md_escape(e['payer_name'])})")
        if e.get("del_req_by") or e.get("edit_req_by"):
            kb.append([InlineKeyboardButton(f"⏳ #{e['id']} — tasdiq kutilmoqda", callback_data=f"delreq_{e['id']}")])
        else:
            kb.append([
                InlineKeyboardButton(f"✏️ #{e['id']}", callback_data=f"editreq_{e['id']}"),
                InlineKeyboardButton(f"🗑 #{e['id']}", callback_data=f"delreq_{e['id']}"),
            ])
    L.append("\n_O'chirish/tahrirlash uchun ikkala investor tasdig'i kerak._")
    await update.message.reply_text("\n".join(L), parse_mode="Markdown",
                                    reply_markup=InlineKeyboardMarkup(kb))


# ---------- O'chirish: ikki investor tasdig'i ----------
@admin_only
async def del_request_cb(update, context):
    """Bir investor o'chirishni so'raydi → ikkinchisiga tasdiq so'rovi boradi."""
    q = update.callback_query
    rid = int(q.data.replace("delreq_",""))
    exp = db.get_expense(rid)
    if not exp:
        await q.answer("Topilmadi"); return
    if exp.get("del_req_by"):
        await q.answer("Bu rasxod uchun so'rov allaqachon yuborilgan.", show_alert=True); return
    db.request_delete(rid, q.from_user.id)
    requester = config.INVESTORS.get(q.from_user.id, "")
    note = exp.get("note", "")
    await q.answer("So'rov yuborildi ✅")
    try:
        await q.edit_message_text(
            f"⏳ #{rid} o'chirish so'raldi ({requester}).\n"
            f"{somm(exp['amount'])} — {note}\n\nIkkinchi investor tasdig'i kutilmoqda.")
    except Exception:
        pass
    text = (f"🗑 *O'chirishga ruxsat so'ralmoqda*\n\n"
            f"{md_escape(requester)} quyidagi rasxodni o'chirmoqchi:\n"
            f"#{rid} {somm(exp['amount'])} — {md_escape(note)}\n\nTasdiqlaysizmi?")
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Ha, o'chirilsin", callback_data=f"delok_{rid}"),
        InlineKeyboardButton("❌ Yo'q", callback_data=f"delno_{rid}"),
    ]])
    photo = exp.get("photo_id")
    for uid in config.ADMIN_IDS:
        if uid != q.from_user.id:
            try:
                if photo:
                    await context.bot.send_photo(uid, photo, caption=text, parse_mode="Markdown", reply_markup=kb)
                else:
                    await context.bot.send_message(uid, text, parse_mode="Markdown", reply_markup=kb)
            except Exception:
                pass


@admin_only
async def del_approve_cb(update, context):
    """Ikkinchi investor tasdiqlaydi → rasxod o'chiriladi. So'ragan o'zi tasdiqlay olmaydi."""
    q = update.callback_query
    rid = int(q.data.replace("delok_",""))
    exp = db.get_expense(rid)
    if not exp:
        await q.answer("Topilmadi")
        try: await q.edit_message_text("❌ Rasxod topilmadi (avval o'chirilgan?).")
        except Exception: pass
        return
    req_by = exp.get("del_req_by")
    if not req_by:
        await q.answer("So'rov bekor qilingan.", show_alert=True); return
    if q.from_user.id == req_by:
        await q.answer("O'zingiz so'ragansiz — ikkinchi investor tasdiqlashi kerak.", show_alert=True); return
    db.delete_expense(rid)
    approver = config.INVESTORS.get(q.from_user.id, "")
    await q.answer("✅ O'chirildi")
    try:
        await q.edit_message_text(f"✅ Rasxod #{rid} o'chirildi (ikkala investor tasdig'i bilan).")
    except Exception:
        pass
    for uid in config.ADMIN_IDS:
        if uid != q.from_user.id:
            try: await context.bot.send_message(uid, f"✅ Rasxod #{rid} o'chirildi ({approver} tasdiqladi).")
            except Exception: pass


@admin_only
async def del_reject_cb(update, context):
    """So'rov rad etiladi — rasxod saqlanib qoladi."""
    q = update.callback_query
    rid = int(q.data.replace("delno_",""))
    db.cancel_delete(rid)
    await q.answer("Bekor qilindi")
    try:
        await q.edit_message_text(f"❌ #{rid} o'chirish rad etildi (rasxod saqlanib qoladi).")
    except Exception:
        pass
    for uid in config.ADMIN_IDS:
        if uid != q.from_user.id:
            try: await context.bot.send_message(uid, f"❌ Rasxod #{rid} o'chirish so'rovi rad etildi.")
            except Exception: pass


# ---------- Tahrirlash: ikki investor tasdig'i ----------
@admin_only
async def edit_request_cb(update, context):
    """Tahrirlashni boshlaydi — yangi summa so'raydi."""
    q = update.callback_query
    rid = int(q.data.replace("editreq_",""))
    exp = db.get_expense(rid)
    if not exp:
        await q.answer("Topilmadi"); return
    if exp.get("edit_req_by") or exp.get("del_req_by"):
        await q.answer("Bu rasxod uchun so'rov allaqachon bor.", show_alert=True); return
    context.user_data.clear()
    context.user_data["flow"] = "ed_amt"
    context.user_data["ed_rid"] = rid
    await q.answer()
    try:
        await q.edit_message_text(
            f"✏️ #{rid} tahrirlash\nEski summa: {somm(exp['amount'])}\n\nYangi summani yozing:")
    except Exception:
        await context.bot.send_message(q.message.chat_id, f"✏️ #{rid} — yangi summani yozing:")


async def edit_text(update, context):
    """ed_amt / ed_note matn bosqichlari."""
    flow = context.user_data.get("flow")
    txt = (update.message.text or "").strip()
    rid = context.user_data.get("ed_rid")
    if flow == "ed_amt":
        amt = parse_amount(txt)
        if not amt or amt <= 0:
            await update.message.reply_text("❌ Summa noto'g'ri. Qaytadan yozing:",
                reply_markup=InlineKeyboardMarkup([[HOME_BTN]]))
            return
        context.user_data["ed_amount"] = amt
        context.user_data["flow"] = "ed_note"
        exp = db.get_expense(rid) or {}
        await update.message.reply_text(
            f"Yangi summa: {somm(amt)}\nEski izoh: {exp.get('note','—')}\n\nYangi izohni yozing:",
            reply_markup=InlineKeyboardMarkup([[HOME_BTN]]))
    elif flow == "ed_note":
        new_note = txt or "—"
        amt = context.user_data.get("ed_amount")
        exp = db.get_expense(rid)
        if not exp:
            context.user_data.clear()
            await update.message.reply_text("❌ Rasxod topilmadi.", reply_markup=BACK_MENU); return
        db.request_edit(rid, update.effective_user.id, amt, new_note)
        requester = config.INVESTORS.get(update.effective_user.id, "")
        context.user_data.clear()
        await update.message.reply_text(
            f"⏳ #{rid} tahrirlash so'raldi. Ikkinchi investor tasdig'i kutilmoqda.",
            reply_markup=BACK_MENU)
        text = (f"✏️ *Tahrirlashga ruxsat so'ralmoqda*\n\n"
                f"{md_escape(requester)} #{rid} ni o'zgartirmoqchi:\n"
                f"Eski: {somm(exp['amount'])} — {md_escape(exp['note'])}\n"
                f"Yangi: {somm(amt)} — {md_escape(new_note)}\n\nTasdiqlaysizmi?")
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Ha", callback_data=f"editok_{rid}"),
            InlineKeyboardButton("❌ Yo'q", callback_data=f"editno_{rid}"),
        ]])
        for uid in config.ADMIN_IDS:
            if uid != update.effective_user.id:
                try: await context.bot.send_message(uid, text, parse_mode="Markdown", reply_markup=kb)
                except Exception: pass


@admin_only
async def edit_approve_cb(update, context):
    q = update.callback_query
    rid = int(q.data.replace("editok_",""))
    exp = db.get_expense(rid)
    if not exp or exp.get("edit_req_by") is None:
        await q.answer("So'rov yo'q", show_alert=True)
        try: await q.edit_message_text("❌ So'rov topilmadi (bekor qilingan?).")
        except Exception: pass
        return
    if q.from_user.id == exp.get("edit_req_by"):
        await q.answer("O'zingiz so'ragansiz — ikkinchi investor tasdiqlashi kerak.", show_alert=True); return
    db.apply_edit(rid)
    approver = config.INVESTORS.get(q.from_user.id, "")
    await q.answer("✅ Saqlandi")
    try:
        await q.edit_message_text(f"✅ Rasxod #{rid} tahrirlandi (ikkala investor tasdig'i bilan).")
    except Exception:
        pass
    for uid in config.ADMIN_IDS:
        if uid != q.from_user.id:
            try: await context.bot.send_message(uid, f"✅ Rasxod #{rid} tahrirlandi ({approver} tasdiqladi).")
            except Exception: pass


@admin_only
async def edit_reject_cb(update, context):
    q = update.callback_query
    rid = int(q.data.replace("editno_",""))
    db.cancel_edit(rid)
    await q.answer("Bekor qilindi")
    try:
        await q.edit_message_text(f"❌ #{rid} tahrirlash rad etildi (eski holicha qoladi).")
    except Exception:
        pass
    for uid in config.ADMIN_IDS:
        if uid != q.from_user.id:
            try: await context.bot.send_message(uid, f"❌ Rasxod #{rid} tahrirlash so'rovi rad etildi.")
            except Exception: pass


# ---------- Tenglashtirish (avtomatik yoziladi, Sozlamalar ichida) ----------
def settle_content():
    hist = db.get_settlements()
    if not hist:
        return "🤝 Tenglashtirish tarixi yo'q (hozircha farq chiqmagan).", None
    L = ["🤝 *Tenglashtirish tarixi*", "_Har oy avtomatik yoziladi._\n"]
    kb = []
    for h in hist:
        mark = "✅ to'langan" if h["paid"] else "⏳ to'lanmagan"
        L.append(f"{h['period']}: {h['payer']} → {h['receiver']} {somm(h['amount'])} [{mark}]")
        if not h["paid"]:
            kb.append([InlineKeyboardButton(f"✅ {h['period']} to'landi deb belgilash",
                                            callback_data=f"setl_{h['period']}")])
    return "\n".join(L), (InlineKeyboardMarkup(kb) if kb else None)


@admin_only
async def tenglashtirish(update, context):
    now = datetime.now(config.TZ)
    finance.auto_settlement(now.year, now.month)
    text, kb = settle_content()
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb or BACK_MENU)


@admin_only
async def open_settle_cb(update, context):
    """Sozlamalar ichidan tenglashtirish tarixini ochadi."""
    q = update.callback_query
    now = datetime.now(config.TZ)
    finance.auto_settlement(now.year, now.month)
    await q.answer()
    text, kb = settle_content()
    try:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    except Exception:
        await context.bot.send_message(q.message.chat_id, text, parse_mode="Markdown", reply_markup=kb)


@admin_only
async def settle_paid_cb(update, context):
    q = update.callback_query
    period = q.data.replace("setl_","")
    db.mark_settlement_paid(period, 1)
    await q.answer("✅ Belgilandi")
    text, kb = settle_content()
    try:
        await q.edit_message_text(text, parse_mode="Markdown", reply_markup=kb)
    except Exception:
        await context.bot.send_message(q.message.chat_id, f"✅ {period} to'langan deb belgilandi.")


# ---------- Sozlamalar (qadoq tannarxlari asosiy botda + tenglashtirish) ----------
@admin_only
async def sozlamalar(update, context):
    quti = store.bd_setting("gift_box_cost", 0)
    oddiy = store.bd_setting("qadoq_oddiy_cost", 0)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🤝 Tenglashtirish tarixi", callback_data="open_settle")],
    ])
    await update.message.reply_text(
        f"⚙️ *Sozlamalar*\n\n"
        f"📦 Oddiy qadoq tannarxi: {somm(oddiy)}\n"
        f"🎁 Sovg'a qutili qadoq tannarxi: {somm(quti)}\n\n"
        f"_Bu tannarxlar *asosiy bot* → ⚙️ Sozlamalar bo'limidan o'rnatiladi va "
        f"har buyurtmaga sotuv paytida biriktiriladi (keyin o'zgartirsangiz, eski oylar o'zgarmaydi)._\n\n"
        f"🤝 Tenglashtirish har oy avtomatik hisoblanib yoziladi.",
        parse_mode="Markdown", reply_markup=kb)


async def maybe_setting_text(update, context):
    """Sozlama uchun kutilayotgan matn (hozircha finance botda sozlama matni yo'q)."""
    return False


# ---------- Hujjat qabul qilish (orders.json / products.json / babydiary.db) ----------
@admin_only
async def document(update, context):
    doc = update.message.document
    name = (doc.file_name or "").lower()
    # --- Backup tiklash: finance.db (moliya bazasi) ---
    if name.endswith(".db") and "finance" in name:
        tmp = os.path.join(tempfile.gettempdir(), f"restore_{doc.file_unique_id}.db")
        f = await context.bot.get_file(doc.file_id)
        await f.download_to_drive(tmp)
        # Haqiqiy finance.db ekanini tekshiramiz
        try:
            con = sqlite3.connect(tmp)
            tbls = {r[0] for r in con.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
            con.close()
        except Exception:
            tbls = set()
        if not {"expenses", "settlements"} <= tbls:
            try: os.remove(tmp)
            except Exception: pass
            await update.message.reply_text(
                "❌ Bu finance.db emas (expenses/settlements jadvali topilmadi).", reply_markup=MENU)
            return
        context.user_data["restore_path"] = tmp
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Ha, tiklash", callback_data="restore_yes"),
            InlineKeyboardButton("❌ Yo'q", callback_data="restore_no")]])
        await update.message.reply_text(
            "♻️ *Moliya bazasini tiklash*\n\nJoriy `finance.db` shu fayl bilan to'liq ALMASHTIRILADI "
            "(barcha rasxod va tenglashtirish tarixi). Ishonchingiz komilmi?",
            parse_mode="Markdown", reply_markup=kb)
        return
    if name.endswith(".db") or "babydiary" in name:
        kind, label = "db", "babydiary.db (savdo statusi)"
    elif "orders" in name:
        kind, label = "orders", "orders.json (buyurtmalar)"
    elif "products" in name:
        kind, label = "products", "products.json (mahsulotlar)"
    else:
        await update.message.reply_text(
            "Bu faylni tanimadim. Kerakli fayllar: orders.json, products.json, babydiary.db",
            reply_markup=MENU)
        return
    # ── Xavfsizlik: jonli fayl allaqachon bor bo'lsa (asosiy bot bilan bir /data) —
    #    uning ustidan YOZMAYMIZ, aks holda production buyurtmalar o'chib ketishi mumkin.
    dst = {"db": config.BABYDIARY_DB, "orders": config.ORDERS_JSON,
           "products": config.PRODUCTS_JSON}[kind]
    if os.path.exists(dst):
        await update.message.reply_text(
            f"ℹ️ {label} allaqachon jonli o'qilyapti (asosiy bot bilan bir bazada).\n"
            "Yuklash shart emas — /holat va Hisobot real-time yangilanadi.\n"
            "(Eski faylni ustidan yozib, ma'lumotni buzib qo'ymaslik uchun qabul qilinmadi.)",
            reply_markup=MENU)
        return
    tmp = os.path.join(tempfile.gettempdir(), doc.file_name)
    f = await context.bot.get_file(doc.file_id)
    await f.download_to_drive(tmp)
    store.save_uploaded(tmp, kind)
    try: os.remove(tmp)
    except Exception: pass
    await update.message.reply_text(f"✅ Qabul qilindi: {label}\n\nEndi /holat yoki 📄 Hisobot yangilanadi.",
                                    reply_markup=MENU)


# ---------- Backup tiklash (restore) tasdiqlash ----------
@admin_only
async def restore_yes_cb(update, context):
    q = update.callback_query
    await q.answer()
    src = context.user_data.pop("restore_path", None)
    if not src or not os.path.exists(src):
        await q.edit_message_text("❌ Tiklash fayli topilmadi. Faylni qaytadan yuklang.")
        return
    import shutil
    try:
        # 1) joriy bazaning WAL'ini yozib yuboramiz va zaxira nusxa olamiz
        try:
            checkpoint_db()
        except Exception:
            pass
        if os.path.exists(config.FINANCE_DB):
            shutil.copy2(config.FINANCE_DB, config.FINANCE_DB + ".pre_restore")
        # 2) eski WAL/SHM sidecar fayllarni o'chiramiz (yangi baza toza ishlasin)
        for ext in ("-wal", "-shm"):
            sc = config.FINANCE_DB + ext
            if os.path.exists(sc):
                try: os.remove(sc)
                except Exception: pass
        # 3) tiklash faylini asosiy bazaning ustiga yozamiz
        shutil.copy2(src, config.FINANCE_DB)
        os.remove(src)
        await q.edit_message_text("✅ *finance.db tiklandi!*\n\n/holat va 📄 Hisobot endi tiklangan ma'lumotni ko'rsatadi.",
                                  parse_mode="Markdown")
    except Exception as e:
        await q.edit_message_text(f"❌ Tiklashda xato: {e}")


@admin_only
async def restore_no_cb(update, context):
    q = update.callback_query
    await q.answer()
    src = context.user_data.pop("restore_path", None)
    if src and os.path.exists(src):
        try: os.remove(src)
        except Exception: pass
    await q.edit_message_text("Bekor qilindi — baza o'zgarmadi.")


async def error_handler(update, context):
    """Kutilmagan xato — asosiy investorni ogohlantiradi.
    O'tkinchi tarmoq xatolarini (Bad Gateway, timeout) JIM o'tkazadi — spam qilmaydi."""
    import traceback
    from telegram.error import NetworkError, TimedOut, RetryAfter, Conflict
    err = getattr(context, "error", None)
    # O'tkinchi Telegram tarmoq xatolari — bot o'zi qayta ulanadi, adminni bezovta qilmaymiz
    if isinstance(err, (NetworkError, TimedOut, RetryAfter, Conflict)):
        print(f"[finance_bot] o'tkinchi xato (e'tiborsiz): {err}")
        return
    tb = "".join(traceback.format_exception_only(type(err), err)).strip() if err else "?"
    print(f"[finance_bot xato] {tb}")
    try:
        await context.bot.send_message(
            5285940949, f"🚨 *Moliya boti xatosi*\n\n`{tb[:500]}`", parse_mode="Markdown")
    except Exception:
        pass


# ---------- 📈 Tahlil ----------
@admin_only
async def tahlil(update, context):
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📦 Sotilmayotgan tovar", callback_data="an_dead")],
        [InlineKeyboardButton("🏆 Tovar foyda reytingi", callback_data="an_profit")],
        [InlineKeyboardButton("👥 Mijozlar tahlili", callback_data="an_cust")],
        [InlineKeyboardButton("🎫 Promo samarasi", callback_data="an_promo")],
        [HOME_BTN],
    ])
    await update.message.reply_text(
        "📈 *Tahlil*\n\nQaysi tahlilni ko'rmoqchisiz?", parse_mode="Markdown", reply_markup=kb)


def _an_back_kb():
    return InlineKeyboardMarkup([[InlineKeyboardButton("◀️ Tahlil", callback_data="an_menu")], [HOME_BTN]])


@admin_only
async def an_menu_cb(update, context):
    q = update.callback_query
    await q.answer()
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📦 Sotilmayotgan tovar", callback_data="an_dead")],
        [InlineKeyboardButton("🏆 Tovar foyda reytingi", callback_data="an_profit")],
        [InlineKeyboardButton("👥 Mijozlar tahlili", callback_data="an_cust")],
        [InlineKeyboardButton("🎫 Promo samarasi", callback_data="an_promo")],
        [HOME_BTN],
    ])
    await q.edit_message_text("📈 *Tahlil*\n\nQaysi tahlilni ko'rmoqchisiz?",
                              parse_mode="Markdown", reply_markup=kb)


@admin_only
async def an_dead_cb(update, context):
    q = update.callback_query
    await q.answer()
    rows = store.dead_stock(30)
    if not rows:
        txt = "📦 *Sotilmayotgan tovar*\n\n✅ 30 kunda sotilmagan, omborda qotgan tovar yo'q."
    else:
        qotgan = sum(r["cost"] * r["stock"] for r in rows)
        L = ["📦 *Sotilmayotgan tovar* (30 kunda sotilmagan)\n"]
        for r in rows[:15]:
            oxir = r["last"] or "hech qachon"
            L.append(f"• {md_escape(r['name'])} — {r['stock']} dona "
                     f"({somm(r['cost']*r['stock'])} qotgan)\n  _oxirgi sotuv: {oxir}_")
        L.append(f"\n💰 Jami qotgan pul: *{somm(qotgan)}*")
        L.append("_Bularga chegirma/aksiya qilsangiz, pul bo'shaydi._")
        txt = "\n".join(L)
    await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=_an_back_kb())


@admin_only
async def an_profit_cb(update, context):
    q = update.callback_query
    await q.answer()
    now = datetime.now(config.TZ)
    top, bottom = finance.product_profit_ranking(now.year, now.month)
    if not top:
        txt = "🏆 *Tovar foyda reytingi*\n\nBu oyda sotuv yo'q."
    else:
        L = [f"🏆 *Tovar foyda reytingi* — {now.year}-{now.month:02d}\n", "*Eng foydali:*"]
        for i, r in enumerate(top, 1):
            L.append(f"{i}. {md_escape(r['name'])} — {somm(r['foyda'])} ({r['qty']} dona)")
        if bottom:
            L.append("\n*Eng kam foyda:*")
            for r in bottom:
                mark = "⚠️ " if r["foyda"] <= 0 else ""
                L.append(f"{mark}{md_escape(r['name'])} — {somm(r['foyda'])} ({r['qty']} dona)")
        txt = "\n".join(L)
    await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=_an_back_kb())


@admin_only
async def an_cust_cb(update, context):
    q = update.callback_query
    await q.answer()
    s = store.customer_stats()
    if not s["customers"]:
        txt = "👥 *Mijozlar tahlili*\n\nHali mijoz yo'q."
    else:
        L = ["👥 *Mijozlar tahlili* (umumiy)\n",
             f"• Jami mijoz: *{s['customers']}*",
             f"• Takroriy mijoz: *{s['repeat']}* ({s['repeat_pct']:.0f}%)",
             f"• O'rtacha chek: *{somm(s['aov'])}*", "\n*Eng ko'p xarid qilganlar:*"]
        for i, c in enumerate(s["top"], 1):
            nm = md_escape(c["name"] or c["phone"])
            L.append(f"{i}. {nm} — {somm(c['total'])} ({c['count']} ta buyurtma)")
        txt = "\n".join(L)
    await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=_an_back_kb())


@admin_only
async def an_promo_cb(update, context):
    q = update.callback_query
    await q.answer()
    rows = store.promo_stats()
    if not rows:
        txt = "🎫 *Promo samarasi*\n\nPromo kod ishlatilgan buyurtma yo'q."
    else:
        L = ["🎫 *Promo samarasi* (umumiy)\n"]
        for r in rows:
            L.append(f"• `{md_escape(r['code'])}` — {r['count']} marta\n"
                     f"  chegirma: {somm(r['discount'])} | savdo: {somm(r['revenue'])}")
        txt = "\n".join(L)
    await q.edit_message_text(txt, parse_mode="Markdown", reply_markup=_an_back_kb())


# ---------- Matn yo'naltirgich ----------
@admin_only
async def text_router(update, context):
    txt = (update.message.text or "").strip()
    if txt in ("🏠 Asosiy menu", "🏠 Bosh menu", "Asosiy menu", "Bosh menu"):
        context.user_data.clear()
        return await go_home(update, context)
    buttons = {
        "➕ Rasxod": rasxod_start, "📊 Holat": holat, "📄 Hisobot (PDF)": hisobot,
        "👤 Balans": balans, "📦 Ombor": ombor,
        "📈 Tahlil": tahlil, "⚙️ Sozlamalar": sozlamalar,
        "✏️ O'zgartirish": ozgartirish,
    }
    # Holat paneli filtr tugmalari (investor bo'yicha)
    _fnames = {f"👤 {n}": uid for uid, n in config.INVESTORS.items()}
    if txt in _fnames:
        return await send_filtered_expenses(update, context, _fnames[txt])
    if txt == "👥 Hammasi":
        return await send_filtered_expenses(update, context, None)
    if txt in buttons:
        context.user_data.pop("flow", None)
        # Holat bo'limida panelga investor filtr tugmalari chiqadi
        menu_kb = filter_menu() if txt == "📊 Holat" else BACK_MENU
        await update.message.reply_text(f"📂 {txt}", reply_markup=menu_kb)
        return await buttons[txt](update, context)
    # sozlama kutilyaptimi
    if await maybe_setting_text(update, context):
        return
    # hisobot — sana oralig'i kiritilyaptimi
    if context.user_data.get("flow") == "rep_custom":
        return await rep_custom_text(update, context)
    # rasxod oqimi
    if context.user_data.get("flow","").startswith("rasxod_"):
        return await rasxod_text(update, context)
    # tahrirlash oqimi (ed_amt / ed_note)
    if context.user_data.get("flow","") in ("ed_amt", "ed_note"):
        return await edit_text(update, context)


def register(app):
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("menu", go_home))
    app.add_handler(CommandHandler("holat", holat))
    app.add_handler(CommandHandler("hisobot", hisobot))
    app.add_handler(CommandHandler("balans", balans))
    app.add_handler(CommandHandler("ombor", ombor))
    app.add_handler(CommandHandler("royxat", royxat))
    app.add_handler(CommandHandler("tenglashtirish", tenglashtirish))
    app.add_handler(CommandHandler("tahlil", tahlil))
    app.add_handler(CallbackQueryHandler(an_menu_cb, pattern="^an_menu$"))
    app.add_handler(CallbackQueryHandler(an_dead_cb, pattern="^an_dead$"))
    app.add_handler(CallbackQueryHandler(an_profit_cb, pattern="^an_profit$"))
    app.add_handler(CallbackQueryHandler(an_cust_cb, pattern="^an_cust$"))
    app.add_handler(CallbackQueryHandler(an_promo_cb, pattern="^an_promo$"))
    # Rasxod turi (rasm bosqichidan keyin) — aniq pattern, rk_norasm bilan to'qnashmasin
    app.add_handler(CallbackQueryHandler(rasxod_kind_cb, pattern="^rk_(tovar|operatsion)$"))
    app.add_handler(CallbackQueryHandler(rasxod_norasm_cb, pattern="^rk_norasm$"))
    app.add_handler(CallbackQueryHandler(rasxod_subcat_cb, pattern="^rsub_"))
    app.add_handler(CallbackQueryHandler(rasxod_save_cb, pattern="^rsave$"))
    # O'chirish — ikki investor tasdig'i
    app.add_handler(CallbackQueryHandler(del_request_cb, pattern="^delreq_"))
    app.add_handler(CallbackQueryHandler(del_approve_cb, pattern="^delok_"))
    app.add_handler(CallbackQueryHandler(del_reject_cb, pattern="^delno_"))
    # Tahrirlash — ikki investor tasdig'i
    app.add_handler(CallbackQueryHandler(edit_request_cb, pattern="^editreq_"))
    app.add_handler(CallbackQueryHandler(edit_approve_cb, pattern="^editok_"))
    app.add_handler(CallbackQueryHandler(edit_reject_cb, pattern="^editno_"))
    # Oy tugmalari (holat / hisobot)
    app.add_handler(CallbackQueryHandler(holat_nav_cb, pattern="^hol_"))
    app.add_handler(CallbackQueryHandler(expense_month_cb, pattern="^exm_"))
    app.add_handler(CallbackQueryHandler(report_dl_cb, pattern="^repdl_"))
    # Hisobot davr tanlash + yuklab olish
    app.add_handler(CallbackQueryHandler(per_today_cb, pattern="^per_today$"))
    app.add_handler(CallbackQueryHandler(per_7_cb, pattern="^per_7$"))
    app.add_handler(CallbackQueryHandler(per_month_cb, pattern="^per_month$"))
    app.add_handler(CallbackQueryHandler(per_back_cb, pattern="^per_back$"))
    app.add_handler(CallbackQueryHandler(per_custom_cb, pattern="^per_custom$"))
    app.add_handler(CallbackQueryHandler(cal_nav_cb, pattern="^caln_"))
    app.add_handler(CallbackQueryHandler(cal_day_cb, pattern="^cald_"))
    app.add_handler(CallbackQueryHandler(cal_nop_cb, pattern="^cal_nop$"))
    app.add_handler(CallbackQueryHandler(rg_pdf_cb, pattern="^rgpdf_"))
    app.add_handler(CallbackQueryHandler(rg_xls_cb, pattern="^rgxls_"))
    app.add_handler(CallbackQueryHandler(restore_yes_cb, pattern="^restore_yes$"))
    app.add_handler(CallbackQueryHandler(restore_no_cb, pattern="^restore_no$"))
    app.add_handler(CallbackQueryHandler(settle_paid_cb, pattern="^setl_"))
    app.add_handler(CallbackQueryHandler(open_settle_cb, pattern="^open_settle$"))
    app.add_handler(CallbackQueryHandler(home_cb, pattern="^home$"))
    app.add_handler(MessageHandler(filters.Document.ALL, document))
    app.add_handler(MessageHandler(filters.PHOTO, rasxod_photo))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_router))


# ===================== jobs.py =====================
"""00:00 vazifalar: finance.db backup + oy almashganda PDF hisobot."""
import os, shutil, tempfile
from datetime import datetime, timedelta


def checkpoint_db():
    """WAL ma'lumotini asosiy .db faylga yozadi (backup to'liq bo'lishi uchun)."""
    try:
        con = sqlite3.connect(config.FINANCE_DB)
        con.execute("PRAGMA wal_checkpoint(TRUNCATE);")
        con.close()
    except Exception as e:
        print(f"[checkpoint] {e}")


async def daily_backup(context):
    if not config.BACKUP_CHANNEL_ID:
        return
    stamp = datetime.now(config.TZ).strftime("%Y-%m-%d")
    if os.path.exists(config.FINANCE_DB):
        checkpoint_db()  # avval WAL'ni asosiy faylga yozamiz
        dst = os.path.join(tempfile.gettempdir(), f"{stamp}_finance.db")
        shutil.copy2(config.FINANCE_DB, dst)
        try:
            with open(dst,"rb") as f:
                await context.bot.send_document(config.BACKUP_CHANNEL_ID, f,
                    filename=os.path.basename(dst), caption=f"🗄 finance.db backup {stamp}")
        finally:
            os.remove(dst)


async def monthly_report(context):
    today = datetime.now(config.TZ).date()
    if today.day != 1:
        return
    prev = today - timedelta(days=1)
    period = f"{prev.year:04d}-{prev.month:02d}"
    if db.report_sent(period):
        return
    r = finance.compute_month(prev.year, prev.month)
    # Tenglashtirishni tarixга yozamiz
    s = r["settlement"]
    if s["amount"] > 0:
        db.save_settlement(period, s["payer"], s["receiver"], s["amount"])
    path = os.path.join(tempfile.gettempdir(), f"BabyDiary_{period}.pdf")
    report_pdf.generate(r, path)
    cap = (f"📄 *{OYLAR[prev.month]} {prev.year}* — oylik hisobot\n"
           f"Sof foyda: {somm(r['sof_foyda'])}")
    if s["amount"] > 0:
        cap += f"\n🤝 {s['payer']} → {s['receiver']} {somm(s['amount'])}"
    ok = False
    for cid in config.REPORT_CHAT_IDS:
        try:
            with open(path,"rb") as f:
                await context.bot.send_document(cid, f, filename=f"BabyDiary_{period}.pdf",
                    caption=cap, parse_mode="Markdown")
            ok = True
        except Exception as e:
            print(f"[monthly_report] {cid}: {e}")
    if ok:
        db.mark_report_sent(period)


# ===================== main.py =====================
"""BabyDiary moliya boti — kirish nuqtasi."""
import logging
from datetime import time
from telegram.ext import Application

logging.basicConfig(format="%(asctime)s [%(levelname)s] %(message)s", level=logging.INFO)
log = logging.getLogger("bdfin")


def main():
    if not config.FINANCE_BOT_TOKEN:
        raise RuntimeError("FINANCE_BOT_TOKEN o'rnatilmagan (Railway env).")
    db.init_db()
    app = Application.builder().token(config.FINANCE_BOT_TOKEN).build()
    handlers.register(app)
    app.add_error_handler(handlers.error_handler)
    midnight = time(hour=0, minute=0, tzinfo=config.TZ)
    app.job_queue.run_daily(jobs.daily_backup, time=midnight, name="backup")
    app.job_queue.run_daily(jobs.monthly_report, time=midnight, name="report")
    log.info("Moliya boti tayyor. Investorlar: %s", list(config.INVESTORS.values()))
    app.run_polling(allowed_updates=["message","callback_query"])


if __name__ == "__main__":
    main()
